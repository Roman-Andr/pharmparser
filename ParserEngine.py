import json
import os
import string
from dataclasses import dataclass
from itertools import chain
from multiprocessing import Pool
from threading import Thread
from typing import List, Tuple, Callable

import psutil
import requests
from bs4 import BeautifulSoup
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from ButtonInjector import run


@dataclass
class Settings:
    __slots__ = ["green", "red", "title", "fileName", "colWidth", "cellWidth", "diffWidth"]
    green: str
    red: str
    title: str
    fileName: str
    colWidth: int
    cellWidth: int
    diffWidth: int


class ParserEngine:
    __slots__ = ["config", "settings", "profiles", "cookies"]

    def __init__(self, config):
        self.config = config
        with open(config, "r", encoding="utf-8") as f:
            loaded = json.load(f)
            self.settings = Settings(**loaded["settings"])
            self.profiles: dict[str, dict[str, str]] = loaded["profiles"]
            self.cookies: dict[str, dict[str, str]] = {"cookies": loaded["cookies"], "data": loaded["data"]}

    def start(self, entries: List[Tuple[str, int]], done: Callable):
        thread = Thread(target=self.process, args=(entries, done))
        thread.start()

    def download(self, target: int, limit: int = 5000):
        psutil.Process(os.getpid()).nice(psutil.HIGH_PRIORITY_CLASS)
        request = requests.post(
            "https://tabletka.by/ajax-request/reload-pharmacy-price",
            headers={
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "X-Requested-With": "XMLHttpRequest",
                "sec-ch-ua-mobile": "?0",
                "Sec-Fetch-Site": "same-origin",
                "Sec-Fetch-Mode": "cors",
                "Sec-Fetch-Dest": "empty",
                "host": "tabletka.by",
            },
            data={
                **self.cookies["data"],
                "id": target,
                "page": "0",
                "sort": "name",
                "sort_type": "asc"
            },
            cookies={
                **self.cookies["cookies"],
                "lim-result": str(limit)
            }
        )
        return json.loads(request.content)["data"]

    def parse(self, file: str):
        psutil.Process(os.getpid()).nice(psutil.HIGH_PRIORITY_CLASS)
        soup = BeautifulSoup(file, 'lxml')
        names = [x.text + ", " + y.text for x, y in zip(
            soup.select("div[class=tooltip-info-header] > a"),
            soup.select("span[class=form-title]")
        )]
        prices = [x.text.strip().rstrip(" р.").lstrip("от ") for x in
                  soup.select("div[class=tooltip-info-header] > span[class=price-value]")]
        entry = {name: float(price) for name, price in zip(names, prices)}
        return entry

    def process(self, entries: List[Tuple[str, int]], done: Callable):
        codes = [y for x, y in entries]
        titles = [x for x, y in entries]
        with Pool(len(codes)) as pool:
            pages = pool.map(self.download, codes)
        with Pool(len(codes)) as pool:
            parse_res = pool.map(self.parse, pages)
        data = dict(zip(codes, parse_res))
        data["titles"] = {y: x for x, y in entries}
        self.excel_export(codes, titles, data)
        done()

    def excel_export(self, codes: List[int], titles: List[str], data):
        offset = 2
        wb = Workbook()
        grid = [*([] for _ in range(offset)),
                ["Название", titles[0]] +
                list(chain(*[[titles[i + 1], "Разница"] for i, x in enumerate(codes) if x != codes[-1]]))]
        names = sorted(list(set(chain(*[list(data[x].keys()) for x in codes]))), key=lambda k: k.lower())
        for x in names:
            prices = []
            for y in codes:
                price1, price2 = data[codes[0]].get(x, "Нет"), data[y].get(x, "Нет")
                prices.append(price2)
                if (price2 == "Нет" or price1 == "Нет") and y != codes[0]:
                    prices.append(0)
                elif y != codes[0]:
                    prices.append(float(f"{float(f'{(float(price2) - float(price1)):.2f}'):+}"))
            row = [x] + prices
            grid.append(row)
        ws = wb.active
        ws.title = self.settings.title
        for x in string.ascii_uppercase:
            ws.column_dimensions[x].width = self.settings.cellWidth
        for x in string.ascii_uppercase[3::2]:
            ws.column_dimensions[x].width = self.settings.diffWidth
        ws.column_dimensions["A"].width = self.settings.colWidth

        for x in string.ascii_uppercase[3::2]:
            red_cell, green_cell = PatternFill(bgColor=self.settings.red), PatternFill(bgColor=self.settings.green)
            dxf_red, dxf_green = DifferentialStyle(fill=red_cell), DifferentialStyle(fill=green_cell)
            rule_less = Rule("cellIs", operator="lessThan", formula=["0"], dxf=dxf_red)
            rule_higher = Rule("cellIs", operator="greaterThan", formula=["0"], dxf=dxf_green)
            [ws.conditional_formatting.add(f"{x}{2 + offset}:{x}{len(grid)}", rule) for rule in (rule_less, rule_higher)]

        ws.auto_filter.ref = f"A{1 + offset}:{get_column_letter(len(grid[0 + offset]))}{len(grid)}"

        [ws.append(x) for x in grid]
        wb.save(self.settings.fileName)
        run(len(grid[0 + offset]))
