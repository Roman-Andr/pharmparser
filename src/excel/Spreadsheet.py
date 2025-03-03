import os
from itertools import chain
from typing import List, Type

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from win32api import RGB

from src.utils import DataType, FilterCriteria, SortOrder, Settings
from .formatters import BaseFormatter
from .macros import Button, ApplyFiltersMacro, SortMacro, RemoveFiltersMacro
from .macros.ButtonInjector import ButtonInjector
from ..utils.utils import remove


class Spreadsheet:
    __slots__ = ["data", "settings", "formatters"]

    def __init__(self, data: DataType, settings: Settings, formatters: List[Type[BaseFormatter]]):
        self.data = data
        self.settings = settings
        self.formatters = formatters

    def export(self, titles: List[str], data: DataType):
        wb = Workbook()
        wb.remove(wb.active)
        for formatter in self.formatters:
            form = formatter(self.settings, data, titles)
            form.format(wb.create_sheet(form.title))
        wb.save(self.settings.fileName)
        column = len(data) * 2
        target = self.settings.fileName.replace('.xlsx', '.xlsm')
        remove(target, os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))

        ButtonInjector(self.settings.fileName, [
            Button('A1', 'Apply Filters',
                   ApplyFiltersMacro(column, FilterCriteria.GREATER_THAN_ZERO),
                   back_color=RGB(18, 230, 89),
                   fore_color=RGB(18, 230, 89)),
            Button('A2', 'Remove Filters',
                   RemoveFiltersMacro(column),
                   back_color=RGB(230, 64, 18),
                   fore_color=RGB(230, 64, 18)),
            *chain(*[[Button(f'{col}1', '↑', SortMacro(col, SortOrder.DESCENDING)),
                      Button(f'{col}2', '↓', SortMacro(col, SortOrder.ASCENDING))]
                     for col in [get_column_letter(x) for x in range(4, column + 2, 2)]])
        ]).save(target)

        remove(self.settings.fileName)
