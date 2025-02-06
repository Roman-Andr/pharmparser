import json
import os
from dataclasses import asdict

from customtkinter import CTk, CTkButton, CTkProgressBar

from ParserEngine import ParserEngine
from Profile import Profile, ProfileSelector


class App(CTk):
    __slots__ = ["progress", "profiles", "current_profile", "engine", "selector", "add_profile_button",
                 "delete_profile_button"]
    processing: bool = False

    def __init__(self):
        super().__init__()

        self.progress: CTkProgressBar = CTkProgressBar(self)
        self.geometry(f"{1100}x{600}")
        self.title("PharmParser")

        self.profiles: list[Profile] = []
        self.current_profile: Profile = None
        CTkButton(self, text="Add", command=self.add_entry).grid(row=1, column=0, padx=30, pady=5)
        CTkButton(self, text="Delete", command=self.delete_entry).grid(row=1, column=1, padx=45, pady=5)
        CTkButton(self, text="Parse", command=self.click).grid(row=1, column=2, padx=30, pady=5)
        self.engine = ParserEngine("config.json")

        for values in self.engine.profiles.values():
            self.profiles.append(Profile(self, values))

        self.selector = ProfileSelector(self, self.profiles)
        self.selector.grid(row=0, column=0, columnspan=3, padx=10, pady=10)

        CTkButton(self, text="Add Profile", command=self.selector.add_profile).grid(row=0, column=3, padx=10, pady=10)
        CTkButton(self, text="Delete Profile", command=self.selector.delete_profile).grid(row=0, column=4, padx=10,
                                                                                          pady=10)

        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def add_entry(self):
        if self.current_profile:
            self.current_profile.add_entry()

    def delete_entry(self):
        if self.current_profile:
            self.current_profile.delete_entry()

    def click(self):
        if self.processing:
            return
        self.processing = True
        self.engine.start(
            [(entry.get_text(), int(entry.get_url().split("/")[-1])) for entry in self.current_profile.entries],
            self.done)
        self.progress.grid(row=sum([len(p.entries) for p in self.profiles]) + 1,
                           column=0,
                           columnspan=3,
                           padx=(20, 10),
                           pady=(10, 10),
                           sticky="ew")
        self.progress.configure(mode="indeterminate")
        self.progress.start()

    def done(self):
        self.processing = False
        self.progress.stop()
        self.progress.grid_forget()
        os.startfile(os.path.abspath(os.getcwd()) + f"\\{self.engine.settings.fileName.replace("xlsx", "xlsm")}")

    def on_closing(self):
        config = {
            "profiles": {},
            "settings": asdict(self.engine.settings),
            "data": self.engine.cookies["data"],
            "cookies": self.engine.cookies["cookies"]
        }

        for i, profile in enumerate(self.profiles):
            config["profiles"][f"Profile {i + 1}"] = {entry.get_text(): entry.get_url() for entry in profile.entries}

        with open(self.engine.config, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=4)
        self.destroy()



import os
from enum import Enum

import pythoncom
import win32com.client as win32
from openpyxl.utils import get_column_letter

from Macro import Macro, ApplyFiltersMacro, RemoveFiltersMacro, SortMacro


class FilterCriteria(Enum):
    GREATER_THAN_ZERO = ">0"
    LESS_THAN_ZERO = "<0"
    GREATER_THAN_OR_EQUAL_ZERO = ">=0"
    LESS_THAN_OR_EQUAL_ZERO = "<=0"
    EQUAL_ZERO = "=0"


class SortOrder(Enum):
    ASCENDING = "xlAscending"
    DESCENDING = "xlDescending"


class ButtonInjector:
    def __init__(self, file_path, *buttons):
        self.file_path = file_path
        pythoncom.CoInitialize()
        self.excel = win32.gencache.EnsureDispatch('Excel.Application')
        self.excel.Visible = False
        self.workbook = None
        self.worksheet = None
        self.buttons = []

        self.open_workbook()
        self.add_buttons(*buttons)

    def open_workbook(self):
        self.workbook = self.excel.Workbooks.Open(os.path.abspath(self.file_path))
        self.worksheet = self.workbook.Sheets(1)

    def close_workbook(self):
        if self.workbook:
            self.workbook.Close()
        self.excel.Quit()

    def save(self, new_file_path):
        self.generate_vba_code()
        if self.workbook:
            self.workbook.SaveAs(os.path.abspath(new_file_path), FileFormat=52)
        self.close_workbook()

    def add_buttons(self, *buttons):
        for button in buttons:
            self.buttons.append(button)

    def generate_vba_code(self):
        for button in self.buttons:
            button.create(self.worksheet)
            if self.workbook:
                module = self.workbook.VBProject.VBComponents.Add(1)
                module.CodeModule.AddFromString(button.macro.get_code().strip())


class Button:
    def __init__(self, cell_address, caption, macro, back_color=None, fore_color=None):
        self.cell_address = cell_address
        self.caption = caption
        self.macro = macro
        self.back_color = back_color
        self.fore_color = fore_color
        self.button_name = None

    def create(self, worksheet):
        cell = worksheet.Range(self.cell_address)
        left = cell.Left
        top = cell.Top
        width = cell.Width
        height = cell.Height

        button = worksheet.Buttons().Add(left, top, width, height)
        button.Caption = self.caption
        button.OnAction = self.macro.name
        self.button_name = button.Name
        self.macro.add_position_code(self.generate_position_code(), self.restore_position_code())

    def generate_position_code(self):
        id_name = self.button_name.replace(' ', '')
        return f"""
        Dim btn{id_name} As Button
        Set btn{id_name} = ActiveSheet.Buttons("{self.button_name}")
        Dim btn{id_name}Left As Double
        Dim btn{id_name}Top As Double
        Dim btn{id_name}Width As Double
        Dim btn{id_name}Height As Double
        btn{id_name}Left = btn{id_name}.Left
        btn{id_name}Top = btn{id_name}.Top
        btn{id_name}Width = btn{id_name}.Width
        btn{id_name}Height = btn{id_name}.Height
        """

    def restore_position_code(self):
        id_name = self.button_name.replace(' ', '')
        return f"""
        btn{id_name}.Left = btn{id_name}Left
        btn{id_name}.Top = btn{id_name}Top
        btn{id_name}.Width = btn{id_name}Width
        btn{id_name}.Height = btn{id_name}Height
        """


def run(column):
    file_path = 'data.xlsx'
    target = 'data.xlsm'

    if os.path.exists(target):
        os.remove(target)

    Macro.start_row, Macro.end_row = 3, 100000
    end_column = column

    buttons = [
        Button('A1', 'Apply Filters',
               ApplyFiltersMacro(end_column, FilterCriteria.GREATER_THAN_ZERO)),
        Button('A2', 'Remove Filters',
               RemoveFiltersMacro(end_column))
    ]
    columns = [get_column_letter(x) for x in range(4, end_column + 2, 2)]
    for col in columns:
        buttons.append(Button(f'{col}2', '↓',
                              SortMacro(col, SortOrder.ASCENDING)))
        buttons.append(Button(f'{col}1', '↑',
                              SortMacro(col, SortOrder.DESCENDING)))

    injector = ButtonInjector(file_path, *buttons)
    injector.save(target)




from customtkinter import CTkEntry


class Entry:
    __slots__ = ["text", "url", "delete_button"]

    def __init__(self, parent, text_placeholder: str = "Pharmacy Name",
                 url_placeholder: str = "https://tabletka.by/pharmacies/****", initial_text: str = "",
                 initial_url: str = ""):
        self.text = self.create_entry(parent, text_placeholder, initial_text)
        self.url = self.create_entry(parent, url_placeholder, initial_url)

    def create_entry(self, parent, placeholder: str, initial_text: str = ""):
        entry = CTkEntry(parent, placeholder_text=placeholder)
        if initial_text:
            entry.insert(0, initial_text)
        entry.bind("<Control-a>", lambda e: ["break", e.widget.select_range(0, 'end'), e.widget.icursor('end')][0])
        entry.bind("<Escape>", lambda e: e.widget.select_clear())
        return entry

    def grid(self, text_row, url_row, column, padx, pady, sticky):
        self.text.grid(row=text_row, column=column, padx=padx, pady=pady, sticky=sticky)
        self.url.grid(row=url_row, column=column + 1, padx=padx, pady=pady, sticky=sticky)

    def destroy(self):
        self.text.grid_forget()
        self.url.grid_forget()

    def get_text(self):
        return self.text.get()

    def get_url(self):
        return self.url.get()




from abc import abstractmethod, ABC


class Macro(ABC):
    start_col = 4
    start_row = 3
    end_row = 100

    def __init__(self, name, code_template):
        self.name = name
        self.code_template = code_template
        self.position_codes = []

    def add_position_code(self, position_code, restore_code):
        self.position_codes.append((position_code, restore_code))

    @abstractmethod
    def get_code(self):
        pass


class ApplyFiltersMacro(Macro):
    def __init__(self, end_column, criteria):
        self.start_column = Macro.start_col
        self.end_column = end_column
        self.criteria = criteria
        code_template = """
        Sub ApplyFilters()
            Application.ScreenUpdating = False
            {position_code_block}
            ActiveSheet.AutoFilterMode = False
            Dim col As Integer
            For col = {start_column} To {end_column} Step 2
                ActiveSheet.Range("{data_range}").AutoFilter Field:=col, Criteria1:="{criteria}"
            Next col
            {restore_code_block}
            Application.ScreenUpdating = True
        End Sub
        """
        super().__init__('ApplyFilters', code_template)

    def get_code(self):
        position_code_block = "\n".join([code for code, _ in self.position_codes])
        restore_code_block = "\n".join([code for _, code in self.position_codes])
        data_range = f"A{Macro.start_row}:T{Macro.end_row}"
        return self.code_template.format(
            position_code_block=position_code_block,
            restore_code_block=restore_code_block,
            start_column=self.start_column,
            end_column=self.end_column,
            criteria=self.criteria.value,
            data_range=data_range
        )


class RemoveFiltersMacro(Macro):
    def __init__(self, end_column):
        self.start_column = Macro.start_col
        self.end_column = end_column
        code_template = """
        Sub RemoveFilters()
            Application.ScreenUpdating = False
            {position_code_block}
            If ActiveSheet.AutoFilterMode Then
                Dim col As Integer
                For col = {start_column} To {end_column} Step 2
                    ActiveSheet.Range("{data_range}").AutoFilter Field:=col
                Next col
            End If
            ActiveSheet.Range("{data_range}").Sort Key1:=ActiveSheet.Columns("A"), Order1:=xlAscending, Header:=xlYes
            {restore_code_block}
            Application.ScreenUpdating = True
        End Sub
        """
        super().__init__('RemoveFilters', code_template)

    def get_code(self):
        position_code_block = "\n".join([code for code, _ in self.position_codes])
        restore_code_block = "\n".join([code for _, code in self.position_codes])
        data_range = f"A{Macro.start_row}:T{Macro.end_row}"
        return self.code_template.format(
            position_code_block=position_code_block,
            restore_code_block=restore_code_block,
            start_column=self.start_column,
            end_column=self.end_column,
            data_range=data_range
        )


class SortMacro(Macro):
    def __init__(self, column, sort_order):
        self.column = column
        self.sort_order = sort_order
        code_template = """
        Sub Sort{sort_name}{column}()
            Application.ScreenUpdating = False
            {position_code_block}
            ActiveSheet.Range("{data_range}").Sort Key1:=ActiveSheet.Columns("{column}"), Order1:={sort_order}, Header:=xlYes
            {restore_code_block}
            Application.ScreenUpdating = True
        End Sub
        """
        super().__init__(f'Sort{sort_order.name}{column}', code_template)

    def get_code(self):
        position_code_block = "\n".join([code for code, _ in self.position_codes])
        restore_code_block = "\n".join([code for _, code in self.position_codes])
        data_range = f"A{Macro.start_row}:T{Macro.end_row}"
        return self.code_template.format(
            position_code_block=position_code_block,
            restore_code_block=restore_code_block,
            column=self.column,
            sort_order=self.sort_order.value,
            sort_name=self.sort_order.name,
            data_range=data_range
        )



from multiprocessing import freeze_support

from customtkinter import set_appearance_mode, set_default_color_theme

from App import App


def main():
    set_appearance_mode("System")
    set_default_color_theme("blue")

    app = App()
    app.mainloop()


if __name__ == "__main__":
    freeze_support()
    main()




import json
import os
import string
from itertools import chain
from multiprocessing import Pool
from threading import Thread
from typing import List, Tuple, Callable

import psutil
import requests
from bs4 import BeautifulSoup
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from ButtonInjector import run
from settings import Settings


class ParserEngine:
    __slots__ = ["config", "settings", "profiles", "cookies"]

    def __init__(self, config):
        self.config = config
        with open(config, "r", encoding="utf-8") as f:
            loaded = json.load(f)
            self.settings = Settings(**loaded["settings"])
            self.profiles: dict[str, dict[str, str]] = loaded["profiles"]
            self.cookies: dict[str, dict[str, str]] = {"cookies": loaded["cookies"], "data": loaded["data"]}

    def start(self, entries: List[Tuple[str, int]], done: Callable):
        thread = Thread(target=self.process, args=(entries, done))
        thread.start()

    def download(self, target: int, limit: int = 5000):
        psutil.Process(os.getpid()).nice(psutil.HIGH_PRIORITY_CLASS)
        request = requests.post(
            "https://tabletka.by/ajax-request/reload-pharmacy-price",
            headers={
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "X-Requested-With": "XMLHttpRequest",
                "sec-ch-ua-mobile": "?0",
                "Sec-Fetch-Site": "same-origin",
                "Sec-Fetch-Mode": "cors",
                "Sec-Fetch-Dest": "empty",
                "host": "tabletka.by",
            },
            data={
                **self.cookies["data"],
                "id": target,
                "page": "0",
                "sort": "name",
                "sort_type": "asc"
            },
            cookies={
                **self.cookies["cookies"],
                "lim-result": str(limit)
            }
        )
        return json.loads(request.content)["data"]

    def parse(self, file: str):
        psutil.Process(os.getpid()).nice(psutil.HIGH_PRIORITY_CLASS)
        soup = BeautifulSoup(file, 'lxml')
        names = [x.text + ", " + y.text for x, y in zip(
            soup.select("div[class=tooltip-info-header] > a"),
            soup.select("span[class=form-title]")
        )]
        prices = [x.text.strip().rstrip(" р.").lstrip("от ") for x in
                  soup.select("div[class=tooltip-info-header] > span[class=price-value]")]
        entry = {name: float(price) for name, price in zip(names, prices)}
        return entry

    def process(self, entries: List[Tuple[str, int]], done: Callable):
        codes = [y for x, y in entries]
        titles = [x for x, y in entries]
        with Pool(len(codes)) as pool:
            pages = pool.map(self.download, codes)
        with Pool(len(codes)) as pool:
            parse_res = pool.map(self.parse, pages)
        data = dict(zip(codes, parse_res))
        data["titles"] = {y: x for x, y in entries}
        self.excel_export(codes, titles, data)
        done()

    def excel_export(self, codes: List[int], titles: List[str], data):
        offset = 2
        wb = Workbook()
        grid = [*([] for _ in range(offset)),
                ["Название", titles[0]] +
                list(chain(*[[titles[i + 1], "Разница"] for i, x in enumerate(codes) if x != codes[-1]]))]
        names = sorted(list(set(chain(*[list(data[x].keys()) for x in codes]))), key=lambda k: k.lower())
        for x in names:
            prices = []
            for y in codes:
                price1, price2 = data[codes[0]].get(x, "Нет"), data[y].get(x, "Нет")
                prices.append(price2)
                if (price2 == "Нет" or price1 == "Нет") and y != codes[0]:
                    prices.append(0)
                elif y != codes[0]:
                    prices.append(float(f"{float(f'{(float(price2) - float(price1)):.2f}'):+}"))
            row = [x] + prices
            grid.append(row)
        ws = wb.active
        ws.title = self.settings.data_title
        for x in string.ascii_uppercase:
            ws.column_dimensions[x].width = self.settings.cellWidth
        for x in string.ascii_uppercase[3::2]:
            ws.column_dimensions[x].width = self.settings.diffWidth
        ws.column_dimensions["A"].width = self.settings.colWidth

        for x in string.ascii_uppercase[3::2]:
            red_cell, green_cell = PatternFill(bgColor=self.settings.red), PatternFill(bgColor=self.settings.green)
            dxf_red, dxf_green = DifferentialStyle(fill=red_cell), DifferentialStyle(fill=green_cell)
            rule_less = Rule("cellIs", operator="lessThan", formula=["0"], dxf=dxf_red)
            rule_higher = Rule("cellIs", operator="greaterThan", formula=["0"], dxf=dxf_green)
            [ws.conditional_formatting.add(f"{x}{2 + offset}:{x}{len(grid)}", rule) for rule in
             (rule_less, rule_higher)]

        ws.auto_filter.ref = f"A{1 + offset}:{get_column_letter(len(grid[0 + offset]))}{len(grid)}"

        [ws.append(x) for x in grid]

        ws = wb.create_sheet(title=self.settings.title)
        grid = [
            ["Асортимент", ],
            ["Позиций ниже всех", ],
        ]

        wb.save(self.settings.fileName)
        run(len(grid[0 + offset]))



from customtkinter import CTkSegmentedButton

from Entry import Entry


class Profile:
    __slots__ = ["parent", "entries"]

    def __init__(self, parent, values):
        self.parent = parent
        self.entries = []
        for title, url in values.items():
            entry = Entry(parent, initial_text=title, initial_url=url)
            self.entries.append(entry)

    def hide(self):
        for entry in self.entries:
            entry.destroy()

    def display(self):
        for i, entry in enumerate(self.entries):
            entry.grid(text_row=i + 2, url_row=i + 2, column=0, padx=(5, 0), pady=(5, 5), sticky="nsew")

    def add_entry(self):
        entry = Entry(self.parent)
        self.entries.append(entry)
        self.display()

    def delete_entry(self):
        if self.entries:
            entry = self.entries.pop()
            entry.destroy()
            self.display()


class ProfileSelector(CTkSegmentedButton):
    __slots__ = ["app", "profiles"]

    def __init__(self, app, profiles, **kwargs):
        super().__init__(app, **kwargs)
        self.app = app
        self.profiles: list[Profile] = profiles
        self.configure(values=[f"Profile {i + 1}" for i in range(len(self.profiles))], command=self.change_profile)
        self.set(f"Profile 1")
        self.change_profile(f"Profile 1")

    def change_profile(self, profile):
        index = int(profile.split(" ")[-1]) - 1
        for p in self.profiles:
            p.hide()
        self.app.current_profile = self.profiles[index]
        self.app.current_profile.display()

    def add_profile(self):
        new_profile_name = f"Profile {len(self.profiles) + 1}"
        self.profiles.append(Profile(self.app, {}))
        self.configure(values=[f"Profile {i + 1}" for i in range(len(self.profiles))])
        self.set(new_profile_name)
        self.change_profile(new_profile_name)

    def delete_profile(self):
        if self.app.current_profile and len(self.profiles) > 1:
            index = self.profiles.index(self.app.current_profile)
            self.app.current_profile.hide()
            self.profiles.pop(index)
            self.configure(values=[f"Profile {i + 1}" for i in range(len(self.profiles))])
            self.set(f"Profile {len(self.profiles)}")
            self.change_profile(f"Profile {len(self.profiles)}")



from dataclasses import dataclass


@dataclass
class Settings:
    __slots__ = ["green", "red", "title", "data_title", "fileName", "colWidth", "cellWidth", "diffWidth"]
    green: str
    red: str
    title: str
    data_title: str
    fileName: str
    colWidth: int
    cellWidth: int
    diffWidth: int


