"""The one place that knows openpyxl.

Everything it writes comes from a :class:`~pharmparser.export.grids.Grid`, so the
report's content and layout are decided — and tested — before Excel is involved.
"""

from __future__ import annotations

import logging
from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .grids import MIN_STYLED_COLUMNS, Grid

logger = logging.getLogger(__name__)

NAVY = "17324D"
TEAL = "159A8C"
PALE_BLUE = "EDF3F8"
PALE_TEAL = "E8F6F3"
PALE_RED = "FDECEC"
INK = "203040"
MUTED = "607385"
WHITE = "FFFFFF"
BORDER_COLOUR = "D8E2EA"


def _apply_column_widths(grid: Grid, ws: Worksheet) -> None:
    """Widths by 1-based column index, indexed with ``get_column_letter``.

    Bug B10 was doing this over ``string.ascii_uppercase``, which silently stopped
    at column Z — around 13 pharmacies.
    """
    for column in range(1, max(grid.width, MIN_STYLED_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(column)].width = grid.column_widths.get(
            column, grid.default_column_width
        )


def _apply_conditional_formatting(grid: Grid, ws: Worksheet) -> None:
    if grid.first_data_row is None or not (grid.below_colour and grid.above_colour):
        return
    below = DifferentialStyle(fill=PatternFill(bgColor=grid.below_colour))
    above = DifferentialStyle(fill=PatternFill(bgColor=grid.above_colour))
    for column in grid.difference_columns:
        letter = get_column_letter(column)
        cells = f"{letter}{grid.first_data_row}:{letter}{grid.last_row}"
        ws.conditional_formatting.add(cells, Rule("cellIs", operator="lessThan", formula=["0"], dxf=below))
        ws.conditional_formatting.add(cells, Rule("cellIs", operator="greaterThan", formula=["0"], dxf=above))


def _style_analysis(grid: Grid, ws: Worksheet) -> None:
    presentation = grid.analysis_presentation
    if presentation is None:
        return

    thin = Side(style="thin", color=BORDER_COLOUR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TEAL
    ws.freeze_panes = f"A{presentation.table_first_row}"
    ws.sheet_view.zoomScale = 90
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:I{grid.last_row}"

    for min_row, min_col, max_row, max_col in presentation.merged_ranges:
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

    title = ws["A1"]
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    title.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 38

    for column in range(1, grid.width + 1):
        cell = ws.cell(2, column)
        cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
        cell.font = Font(name="Aptos", size=11, bold=column in {1, 9}, color=INK)
        cell.alignment = Alignment(vertical="center", horizontal="right" if column == 9 else "left")
    ws.row_dimensions[2].height = 28

    for row in presentation.section_rows:
        cell = ws.cell(row, 1)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name="Aptos Display", size=12, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 25

    for row in presentation.metric_rows:
        ws.row_dimensions[row].height = 38
        for label_column in (1, 3, 5, 7):
            label = ws.cell(row, label_column)
            value = ws.cell(row, label_column + 1)
            if label.value is None:
                continue
            for cell in (label, value):
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            label.fill = PatternFill("solid", fgColor=PALE_BLUE)
            label.font = Font(name="Aptos", size=10, color=MUTED)
            value.fill = PatternFill("solid", fgColor=WHITE)
            value.font = Font(name="Aptos Display", size=15, bold=True, color=NAVY)
            value.alignment = Alignment(horizontal="center", vertical="center")

    header_row = presentation.table_header_row
    ws.row_dimensions[header_row].height = 38
    for cell in ws[header_row][: grid.width]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if presentation.table_last_row >= presentation.table_first_row:
        ws.auto_filter.ref = f"A{header_row}:I{presentation.table_last_row}"
        for row in range(presentation.table_first_row, presentation.table_last_row + 1):
            ws.row_dimensions[row].height = 24
            fill = WHITE if (row - presentation.table_first_row) % 2 == 0 else PALE_BLUE
            for column in range(1, grid.width + 1):
                cell = ws.cell(row, column)
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(name="Aptos", size=10, color=INK)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left" if column == 1 else "center",
                    vertical="center",
                    wrap_text=column == 1,
                )

        difference_range = f"I{presentation.table_first_row}:I{presentation.table_last_row}"
        ws.conditional_formatting.add(
            difference_range,
            Rule(
                "cellIs",
                operator="greaterThan",
                formula=["0"],
                dxf=DifferentialStyle(fill=PatternFill("solid", fgColor=PALE_TEAL)),
            ),
        )
        ws.conditional_formatting.add(
            difference_range,
            Rule(
                "cellIs",
                operator="lessThan",
                formula=["0"],
                dxf=DifferentialStyle(fill=PatternFill("solid", fgColor=PALE_RED)),
            ),
        )

    for row in presentation.note_rows:
        cell = ws.cell(row, 1)
        cell.font = Font(name="Aptos", size=9, italic=True, color=MUTED)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 28

    for cell in ("B5", "D5", "F5", "B6", "F6", "H6", "B9", "D9", "F9", "H9"):
        ws[cell].number_format = "0"
    ws["H5"].number_format = '0.00 "BYN"'
    ws["D6"].number_format = "0.00"
    ws["B10"].number_format = '0.00 "BYN"'
    for cell in ("D10", "F10"):
        ws[cell].number_format = '0.00"%"'
    for row in range(presentation.table_first_row, presentation.table_last_row + 1):
        ws.cell(row, 8).number_format = '0.00 "BYN"'
        ws.cell(row, 9).number_format = '0.00"%"'


def render(grid: Grid, ws: Worksheet) -> None:
    """Write one grid into an existing worksheet."""
    _apply_column_widths(grid, ws)
    _apply_conditional_formatting(grid, ws)
    if grid.header_row is not None:
        ws.auto_filter.ref = f"A{grid.header_row}:{get_column_letter(grid.width)}{grid.last_row}"
    for row in grid.rows:
        ws.append(list(row))
    _style_analysis(grid, ws)


def build_workbook(grids: Sequence[Grid], title: str | None = None) -> Workbook:
    """An in-memory workbook holding every grid, one sheet each."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    if title:
        workbook.properties.title = title
    for grid in grids:
        render(grid, workbook.create_sheet(grid.title))
    return workbook


def write_grids(grids: Sequence[Grid], path: Path, title: str | None = None) -> Path:
    """Save every grid to ``path`` as a plain ``.xlsx``.

    Needs neither Excel nor Windows, which is what makes the CLI and the
    integration tests possible.
    """
    build_workbook(grids, title).save(path)
    logger.info("Wrote %s", path)
    return path
