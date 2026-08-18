from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ...config import ExportSettings
from ...domain import DifferenceFn, PriceTable, comparison_rows
from .base_formatter import BaseFormatter

HEADER_OFFSET = 2
"""Blank spacer rows above the header, left free for the macro buttons."""

NOT_STOCKED = "Нет"
"""Displayed where a pharmacy does not stock an item.

Only a rendering concern: the domain represents absence as ``None`` so it never
travels inside otherwise-numeric data (B9).
"""

Cell = str | float | None


class DataFormatter(BaseFormatter):
    __slots__ = ["difference"]

    def __init__(self, settings: ExportSettings, table: PriceTable, difference: DifferenceFn):
        super().__init__(settings, table)
        self.difference = difference

    @property
    def _total_columns(self) -> int:
        # "Название" + the reference price, then a price and a difference per competitor.
        return 2 + 2 * len(self.table.competitors)

    @property
    def _difference_columns(self) -> list[int]:
        return list(range(4, self._total_columns + 1, 2))

    def _header(self) -> list[Cell]:
        header: list[Cell] = ["Название", self.table.reference.name]
        for competitor in self.table.competitors:
            header += [competitor.name, "Разница"]
        return header

    def _rows(self) -> list[list[Cell]]:
        rows: list[list[Cell]] = []
        for row in comparison_rows(self.table, self.difference):
            cells: list[Cell] = [row.item, row.prices[0] if row.prices[0] is not None else NOT_STOCKED]
            for price, difference in zip(row.prices[1:], row.differences, strict=True):
                cells.append(price if price is not None else NOT_STOCKED)
                # Left blank when undefined, so it stays distinguishable from a
                # genuine difference of zero.
                cells.append(difference)
            rows.append(cells)
        return rows

    def _apply_conditional_formatting(self, ws: Worksheet, last_row: int) -> None:
        dxf_red = DifferentialStyle(fill=PatternFill(bgColor=self.settings.red))
        dxf_green = DifferentialStyle(fill=PatternFill(bgColor=self.settings.green))
        for column in self._difference_columns:
            letter = get_column_letter(column)
            cells = f"{letter}{HEADER_OFFSET + 2}:{letter}{last_row}"
            ws.conditional_formatting.add(cells, Rule("cellIs", operator="lessThan", formula=["0"], dxf=dxf_red))
            ws.conditional_formatting.add(
                cells, Rule("cellIs", operator="greaterThan", formula=["0"], dxf=dxf_green)
            )

    def format(self, ws: Worksheet) -> None:
        grid: list[list[Cell]] = [*([] for _ in range(HEADER_OFFSET)), self._header(), *self._rows()]
        last_row = len(grid)

        widths = {1: self.settings.col_width}
        for column in self._difference_columns:
            widths[column] = self.settings.diff_width
        self._set_column_widths(ws, widths, self.settings.cell_width, self._total_columns)

        self._apply_conditional_formatting(ws, last_row)
        ws.auto_filter.ref = f"A{HEADER_OFFSET + 1}:{get_column_letter(self._total_columns)}{last_row}"

        for row in grid:
            ws.append(row)
