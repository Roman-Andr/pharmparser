import os
from itertools import chain

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from ..config import ExportSettings
from ..domain import PriceTable
from ..utils import FilterCriteria, SortOrder
from ..utils.file_utils import clean_temp_files, remove
from .formatters import BaseFormatter, DataFormatter
from .macros import ApplyFiltersMacro, Button, RemoveFiltersMacro, SortMacro


class ExcelManager:
    """Scoped Excel COM application.

    The pywin32 modules are imported lazily so that importing this module does not
    require Windows; only entering the context manager does.
    """

    def __init__(self):
        self.excel = None
        self._pythoncom = None

    def __enter__(self):
        import pythoncom
        import win32com.client as win32

        self._pythoncom = pythoncom
        pythoncom.CoInitialize()
        self.excel = win32.Dispatch('Excel.Application')
        return self.excel

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.excel.Quit()
        self._pythoncom.CoUninitialize()


class Spreadsheet:
    __slots__ = ["formatters", "settings", "table"]

    def __init__(
        self, table: PriceTable, settings: ExportSettings, formatters: list[tuple[BaseFormatter, str]]
    ):
        self.table = table
        self.settings = settings
        self.formatters = formatters

    def export(self) -> str:
        """Write the workbook and return the path of the .xlsm actually produced."""
        from win32api import RGB

        with ExcelManager() as excel:
            try:
                for workbook in excel.Workbooks:
                    if workbook.FullName == os.path.abspath(self.settings.macro_file_name):
                        workbook.Close(SaveChanges=False)
                        break
            except Exception:
                pass

        wb = Workbook()
        wb.remove(wb.active)
        end_column = len(self.table.pharmacies) * 2
        target = self.settings.macro_file_name
        clean_temp_files(target)
        sheet_titles = []
        for formatter, title in self.formatters:
            sheet = wb.create_sheet(title)
            formatter.format(sheet)
            if isinstance(formatter, DataFormatter):
                sheet_titles.append(title)
        wb.save(self.settings.file_name)
        for i, sheet_name in enumerate(sheet_titles):
            with ExcelManager() as excel:
                inject(excel, i + 1, self.settings.file_name if i == 0 else f"{i - 1}{target}", [
                    Button('A1', 'Apply Filters',
                           ApplyFiltersMacro(end_column, FilterCriteria.GREATER_THAN_ZERO, sheet_name),
                           back_color=RGB(18, 230, 89),
                           fore_color=RGB(18, 230, 89)),
                    Button('A2', 'Remove Filters',
                           RemoveFiltersMacro(end_column, sheet_name),
                           back_color=RGB(230, 64, 18),
                           fore_color=RGB(230, 64, 18)),
                    *chain(
                        *[[Button(f'{col}1', '↑', SortMacro(col, end_column, SortOrder.DESCENDING, sheet_name)),
                           Button(f'{col}2', '↓', SortMacro(col, end_column, SortOrder.ASCENDING, sheet_name))]
                          for col in [get_column_letter(x) for x in range(4, end_column + 2, 2)]])
                ], f"{i}{target}")
            remove(f"{i - 1}{target}")
        os.replace(f"{len(sheet_titles) - 1}{target}", target)
        remove(self.settings.file_name)
        return target


def inject(excel, btn_id, file_path, buttons, new_file_path):
    excel.Visible = False
    workbook = excel.Workbooks.Open(os.path.abspath(file_path))
    worksheet = workbook.Sheets(btn_id)
    sheet_name = worksheet.Name
    for button in buttons:
        button.macro.sheet_name = sheet_name
        button.create(worksheet)
        if workbook:
            module = workbook.VBProject.VBComponents.Add(1)
            module.CodeModule.AddFromString(button.macro.code.strip())
    if workbook:
        workbook.SaveAs(os.path.abspath(new_file_path), FileFormat=52)
        workbook.Close()
