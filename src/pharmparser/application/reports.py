"""Five-sheet report generation from immutable run history."""

from __future__ import annotations

import os
import tempfile
from decimal import Decimal
from pathlib import Path
from statistics import median
from uuid import UUID

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from ..domain import Product, RunStatus, may_export
from ..export.vba.ovba import build_project
from ..export.vba.xlsm import ButtonSpec, package
from .history import HistoryRepository
from .settings import DesktopSettings

SHEETS = ("Обзор", "Сравнение", "Изменения", "История", "Проблемы")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")


class ReportService:
    def __init__(self, history: HistoryRepository) -> None:
        self.history = history

    def export(
        self,
        run_id: UUID,
        settings: DesktopSettings,
        *,
        path: Path | None = None,
        format_: str | None = None,
    ) -> Path:
        run = self.history.get_run(run_id)
        if not may_export(run.status):
            raise ValueError("экспорт доступен только для полного или неполного успешного запуска")
        chosen_format = format_ or settings.report_format
        if chosen_format not in {"xlsx", "xlsm"}:
            raise ValueError("формат должен быть xlsx или xlsm")
        target = path or self._next_path(run_id, settings, chosen_format)
        target = target.with_suffix(f".{chosen_format}").absolute()
        target.parent.mkdir(parents=True, exist_ok=True)

        workbook = self.build(run_id, settings)
        with tempfile.TemporaryDirectory(dir=target.parent, prefix=".pharmparser-") as temporary:
            plain = Path(temporary) / "report.xlsx"
            workbook.save(plain)
            if chosen_format == "xlsx":
                os.replace(plain, target)
            else:
                built = Path(temporary) / "report.xlsm"
                source, buttons = _macro_project(workbook["Сравнение"], workbook["Изменения"])
                project = build_project({"PharmParser": source}, list(workbook.sheetnames))
                package(plain, built, project, buttons, workbook.sheetnames)
                os.replace(built, target)
        self.history.add_artifact(run_id, chosen_format, target)
        return target

    def _next_path(self, run_id: UUID, settings: DesktopSettings, format_: str) -> Path:
        run = self.history.get_run(run_id)
        profile = self.history.profile_snapshot(run_id)
        timestamp = run.started_at.astimezone()
        values = {
            "profile": _safe_name(profile.name),
            "date": timestamp.strftime("%Y-%m-%d"),
            "time": timestamp.strftime("%H-%M"),
        }
        try:
            stem = settings.file_name_template.format_map(values)
        except (KeyError, ValueError) as error:
            raise ValueError(f"неверный шаблон имени файла: {error}") from error
        directory = Path(settings.output_directory)
        candidate = directory / f"{stem}.{format_}"
        index = 2
        while candidate.exists():
            candidate = directory / f"{stem}_{index}.{format_}"
            index += 1
        return candidate

    def build(self, run_id: UUID, settings: DesktopSettings) -> Workbook:
        profile = self.history.profile_snapshot(run_id)
        prices = self.history.prices_for_run(run_id)
        previous_id = self.history.previous_completed(run_id)
        previous = self.history.prices_for_run(previous_id) if previous_id else {}

        workbook = Workbook()
        workbook.remove(workbook.active)
        workbook.properties.title = f"Сравнение цен — {profile.name}"
        self._overview(workbook.create_sheet("Обзор"), run_id, prices)
        self._comparison(workbook.create_sheet("Сравнение"), run_id, prices, settings)
        self._changes(workbook.create_sheet("Изменения"), prices, previous)
        self._history(workbook.create_sheet("История"), profile.id, profile.reference_pharmacy_id)
        self._problems(workbook.create_sheet("Проблемы"), run_id)
        return workbook

    def _overview(self, sheet: Worksheet, run_id: UUID, prices: dict[str, dict[Product, Decimal]]) -> None:
        run = self.history.get_run(run_id)
        profile = self.history.profile_snapshot(run_id)
        reference = next(item for item in profile.pharmacies if item.id == run.reference_pharmacy_id)
        sheet.append(["Обзор отчета"])
        sheet["A1"].font = Font(size=18, bold=True)
        if run.status is RunStatus.PARTIAL:
            sheet.append(["ВНИМАНИЕ: отчет неполный — часть аптек не загрузилась"])
            sheet["A2"].fill = WARNING_FILL
            sheet["A2"].font = Font(bold=True, color="9C6500")
        sheet.append(["Профиль", profile.name])
        sheet.append(["Время запуска", run.started_at.replace(tzinfo=None)])
        sheet.append(["Полнота", "Полный" if run.status is RunStatus.COMPLETED else "Неполный"])
        sheet.append(["Основная аптека", reference.name])
        sheet.append(["Аптек загружено", run.successful_pharmacies, "из", run.pharmacy_count])
        sheet.append(["Уникальных товаров", len({product for values in prices.values() for product in values})])

        reference_prices = prices.get(reference.id, {})
        sheet.append([])
        header_row = sheet.max_row + 1
        sheet.append(["Аптека", "Ассортимент", "Дешевле основной", "Равно", "Дороже основной"])
        for entry in profile.pharmacies:
            pharmacy_prices = prices.get(entry.id, {})
            shared = set(reference_prices) & set(pharmacy_prices)
            sheet.append(
                [
                    entry.name,
                    len(pharmacy_prices),
                    sum(pharmacy_prices[item] < reference_prices[item] for item in shared),
                    sum(pharmacy_prices[item] == reference_prices[item] for item in shared),
                    sum(pharmacy_prices[item] > reference_prices[item] for item in shared),
                ]
            )
        _style_header(sheet, header_row)
        sheet.column_dimensions["A"].width = 44
        sheet.column_dimensions["B"].width = 24

    def _comparison(
        self,
        sheet: Worksheet,
        run_id: UUID,
        prices: dict[str, dict[Product, Decimal]],
        settings: DesktopSettings,
    ) -> None:
        run = self.history.get_run(run_id)
        profile = self.history.profile_snapshot(run_id)
        reference = next(item for item in profile.pharmacies if item.id == run.reference_pharmacy_id)
        competitors = [item for item in profile.pharmacies if item.id != reference.id]
        header = ["Название", "Форма и дозировка", "Производитель", f"Цена — {reference.name}"]
        for entry in competitors:
            header.extend([f"Цена — {entry.name}", f"Разница BYN — {entry.name}", f"Разница % — {entry.name}"])
        header.extend(["_Наша цена ниже всех", "_Есть конкурент дешевле"])
        sheet.append(header)
        all_products = sorted({product for value in prices.values() for product in value}, key=lambda p: p.key)
        reference_prices = prices.get(reference.id, {})
        for product in all_products:
            base = reference_prices.get(product)
            row: list[object] = [product.name, product.form, product.manufacturer, _excel_money(base)]
            comparisons: list[int] = []
            for entry in competitors:
                other = prices.get(entry.id, {}).get(product)
                difference = other - base if base is not None and other is not None else None
                percent = difference / base if difference is not None and base else None
                row.extend(
                    [_excel_money(other), _excel_money(difference), float(percent) if percent is not None else None]
                )
                if difference is not None:
                    comparisons.append((difference > 0) - (difference < 0))
            row.extend([
                "Yes" if comparisons and all(value >= 0 for value in comparisons) else "No",
                "Yes" if any(value < 0 for value in comparisons) else "No",
            ])
            sheet.append(row)
        sheet.insert_rows(1, amount=2)
        _make_table(sheet, "ComparisonTable", header_row=3)
        sheet.freeze_panes = "D4"
        sheet.column_dimensions["A"].width = 32
        sheet.column_dimensions["B"].width = 34
        sheet.column_dimensions["C"].width = 28
        sheet.column_dimensions[get_column_letter(len(header) - 1)].hidden = True
        sheet.column_dimensions[get_column_letter(len(header))].hidden = True
        red = PatternFill("solid", fgColor=settings.red)
        green = PatternFill("solid", fgColor=settings.green)
        for column in range(6, len(header) - 1, 3):
            cells = f"{get_column_letter(column)}4:{get_column_letter(column)}{sheet.max_row}"
            sheet.conditional_formatting.add(cells, CellIsRule(operator="lessThan", formula=["0"], fill=red))
            sheet.conditional_formatting.add(cells, CellIsRule(operator="greaterThan", formula=["0"], fill=green))
        for row in sheet.iter_rows(min_row=4, min_col=4, max_col=len(header) - 2):
            for raw_cell in row:
                cell = raw_cell
                assert isinstance(cell, Cell)
                if (cell.column - 4) % 3 in {0, 1}:
                    cell.number_format = '#,##0.00 "BYN"'
                elif (cell.column - 4) % 3 == 2:
                    cell.number_format = "0.00%"

    def _changes(
        self,
        sheet: Worksheet,
        current: dict[str, dict[Product, Decimal]],
        previous: dict[str, dict[Product, Decimal]],
    ) -> None:
        sheet.append(
            [
                "Аптека",
                "Название",
                "Форма и дозировка",
                "Производитель",
                "Предыдущая цена",
                "Текущая цена",
                "Изменение BYN",
                "Изменение %",
                "Статус",
            ]
        )
        pharmacy_ids = sorted(set(current) | set(previous))
        for pharmacy_id in pharmacy_ids:
            old_prices = previous.get(pharmacy_id, {})
            new_prices = current.get(pharmacy_id, {})
            for product in sorted(set(old_prices) | set(new_prices), key=lambda item: item.key):
                old = old_prices.get(product)
                new = new_prices.get(product)
                difference = new - old if old is not None and new is not None else None
                percent = difference / old if difference is not None and old else None
                if old is None:
                    status = "Новая"
                elif new is None:
                    status = "Исчезла"
                else:
                    status = "Изменилась" if difference else "Без изменений"
                sheet.append([
                    pharmacy_id, product.name, product.form, product.manufacturer,
                    _excel_money(old), _excel_money(new), _excel_money(difference),
                    float(percent) if percent is not None else None, status,
                ])
        sheet.insert_rows(1, amount=2)
        _make_table(sheet, "ChangesTable", header_row=3)
        sheet.freeze_panes = "B4"
        for column in (5, 6, 7):
            for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=4):
                for item in cell:
                    item.number_format = '#,##0.00 "BYN"'
        for cell in sheet["H"][3:]:
            cell.number_format = "0.00%"

    def _history(self, sheet: Worksheet, profile_id: UUID, reference_id: str | None) -> None:
        sheet.append(["Дата", "Аптека", "Ассортимент", "Медианный индекс", "Дешевле", "Равно", "Дороже"])
        for run in reversed(self.history.list_runs(profile_id)):
            if run.status is not RunStatus.COMPLETED or run.reference_pharmacy_id != reference_id:
                continue
            profile = self.history.profile_snapshot(run.id)
            prices = self.history.prices_for_run(run.id)
            reference_prices = prices.get(run.reference_pharmacy_id, {})
            for pharmacy in profile.pharmacies:
                if pharmacy.id == run.reference_pharmacy_id:
                    continue
                values = prices.get(pharmacy.id, {})
                shared = set(reference_prices) & set(values)
                indexes = [float(values[item] / reference_prices[item]) for item in shared if reference_prices[item]]
                sheet.append([
                    run.started_at.replace(tzinfo=None), pharmacy.name, len(values),
                    median(indexes) if indexes else None,
                    sum(values[item] < reference_prices[item] for item in shared),
                    sum(values[item] == reference_prices[item] for item in shared),
                    sum(values[item] > reference_prices[item] for item in shared),
                ])
        _make_table(sheet, "HistoryTable")
        if sheet.max_row > 1:
            chart = LineChart()
            chart.title = "Динамика медианного индекса цен"
            chart.y_axis.title = "Индекс к основной аптеке"
            chart.add_data(Reference(sheet, min_col=4, min_row=1, max_row=sheet.max_row), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row))
            sheet.add_chart(chart, "I2")

    def _problems(self, sheet: Worksheet, run_id: UUID) -> None:
        sheet.append(["Время", "Аптека", "Код", "Описание"])
        for warning in self.history.warnings_for_run(run_id):
            sheet.append([warning["created_at"], warning["pharmacy_id"], warning["code"], warning["message"]])
        for attempt in self.history.attempts_for_run(run_id):
            if attempt["status"] != "completed":
                sheet.append([
                    attempt["finished_at"] or attempt["started_at"], attempt["pharmacy_id"],
                    attempt["error_code"] or attempt["status"], attempt["error_message"] or "Аптека пропущена",
                ])
        _make_table(sheet, "ProblemsTable")
        sheet.column_dimensions["D"].width = 80


def _make_table(sheet: Worksheet, name: str, *, header_row: int = 1) -> None:
    _style_header(sheet, header_row)
    if sheet.max_row == header_row:
        sheet.append([None] * sheet.max_column)
    table = Table(
        displayName=name,
        ref=f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)
    sheet.auto_filter.ref = table.ref


def _style_header(sheet: Worksheet, row: int) -> None:
    for cell in sheet[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _excel_money(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _safe_name(value: str) -> str:
    cleaned = "".join("_" if character in '<>:"/\\|?*' else character for character in value).strip(" .")
    return cleaned or "Профиль"


def _macro_project(comparison: Worksheet, changes: Worksheet) -> tuple[str, dict[str, list[ButtonSpec]]]:
    competitor_count = max(0, (comparison.max_column - 6) // 3)
    cheapest_field = comparison.max_column - 1
    competitor_field = comparison.max_column
    parts = [
        "Option Explicit",
        "Private Function PPTable() As ListObject",
        "Set PPTable = ActiveSheet.ListObjects(1)",
        "End Function",
        "Sub ReferenceCheapest()",
        f'PPTable.Range.AutoFilter Field:={cheapest_field}, Criteria1:="Yes"',
        "End Sub",
        "Sub CompetitorCheaper()",
        f'PPTable.Range.AutoFilter Field:={competitor_field}, Criteria1:="Yes"',
        "End Sub",
        "Sub ResetFilters()",
        "On Error Resume Next",
        "ActiveSheet.ShowAllData",
        "On Error GoTo 0",
        "End Sub",
    ]
    buttons = {
        "Сравнение": [
            ButtonSpec("A1", "Наша цена ниже всех", "ReferenceCheapest"),
            ButtonSpec("B1", "Есть конкурент дешевле", "CompetitorCheaper"),
            ButtonSpec("C1", "Сбросить фильтры", "ResetFilters"),
        ],
        "Изменения": [],
    }
    for index in range(competitor_count):
        field = 6 + index * 3
        up = f"SortComparisonUp{index + 1}"
        down = f"SortComparisonDown{index + 1}"
        parts.extend(_sort_macro(up, "ComparisonTable", field, "xlAscending"))
        parts.extend(_sort_macro(down, "ComparisonTable", field, "xlDescending"))
        column = get_column_letter(field)
        buttons["Сравнение"].extend([
            ButtonSpec(f"{column}1", "↑", up),
            ButtonSpec(f"{column}2", "↓", down),
        ])
    parts.extend(_sort_macro("SortChangesUp", "ChangesTable", 7, "xlAscending"))
    parts.extend(_sort_macro("SortChangesDown", "ChangesTable", 7, "xlDescending"))
    buttons["Изменения"] = [
        ButtonSpec("G1", "↑", "SortChangesUp"),
        ButtonSpec("G2", "↓", "SortChangesDown"),
    ]
    return "\n".join(parts) + "\n", buttons


def _sort_macro(name: str, table: str, field: int, order: str) -> list[str]:
    return [
        f"Sub {name}()",
        "Dim lo As ListObject",
        f'Set lo = ActiveSheet.ListObjects("{table}")',
        "lo.Sort.SortFields.Clear",
        f"lo.Sort.SortFields.Add Key:=lo.ListColumns({field}).Range, SortOn:=xlSortOnValues, Order:={order}",
        "With lo.Sort",
        ".Header = xlYes",
        ".Apply",
        "End With",
        "End Sub",
    ]
