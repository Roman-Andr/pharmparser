import os
from itertools import chain
from typing import List, Tuple

import pythoncom
import win32com.client as win32
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from win32api import RGB

from utils import DataType, FilterCriteria, Settings, SortOrder
from utils.file_utils import clean_temp_files, remove
from .formatters import BaseFormatter
from .formatters import DataFormatter
from .macros import Button, ApplyFiltersMacro, RemoveFiltersMacro, SortMacro


class ExcelManager:
    def __init__(self):
        self.excel = None

    def __enter__(self):
        pythoncom.CoInitialize()
        self.excel = win32.Dispatch('Excel.Application')
        return self.excel

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.excel.Quit()
        pythoncom.CoUninitialize()


class Spreadsheet:
    __slots__ = ["data", "settings", "formatters"]

    def __init__(self, data: DataType, settings: Settings, formatters: List[Tuple[BaseFormatter, str]]):
        self.data = data
        self.settings = settings
        self.formatters = formatters

    def export(self, data: DataType):
        with ExcelManager() as excel:
            try:
                for workbook in excel.Workbooks:
                    if workbook.FullName == os.path.abspath(self.settings.fileName.replace('.xlsx', '.xlsm')):
                        workbook.Close(SaveChanges=False)
                        break
            except Exception:
                pass

        wb = Workbook()
        wb.remove(wb.active)
        end_column = len(data) * 2
        target = self.settings.fileName.replace('.xlsx', '.xlsm')
        clean_temp_files(target)
        sheet_titles = []
        for formatter, title in self.formatters:
            sheet = wb.create_sheet(title)
            formatter.format(sheet)
            if isinstance(formatter, DataFormatter):
                sheet_titles.append(title)
        wb.save(self.settings.fileName)
        for i, sheet_name in enumerate(sheet_titles):
            with ExcelManager() as excel:
                inject(excel, i + 1, self.settings.fileName if i == 0 else f"{i - 1}{target}", [
                    Button('A1', 'Apply Filters',
                           ApplyFiltersMacro(end_column, FilterCriteria.GREATER_THAN_ZERO, sheet_name),
                           back_color=RGB(18, 230, 89),
                           fore_color=RGB(18, 230, 89)),
                    Button('A2', 'Remove Filters',
                           RemoveFiltersMacro(end_column, sheet_name),
                           back_color=RGB(230, 64, 18),
                           fore_color=RGB(230, 64, 18)),
                    *chain(
                        *[[Button(f'{col}1', '↑', SortMacro(col, end_column, SortOrder.DESCENDING, sheet_name)),
                           Button(f'{col}2', '↓', SortMacro(col, end_column, SortOrder.ASCENDING, sheet_name))]
                          for col in [get_column_letter(x) for x in range(4, end_column + 2, 2)]])
                ], f"{i}{target}")
            remove(f"{i - 1}{target}")
        os.rename(f"{len(sheet_titles) - 1}{target}", target)
        remove(self.settings.fileName)


def inject(excel, btn_id, file_path, buttons, new_file_path):
    excel.Visible = False
    workbook = excel.Workbooks.Open(os.path.abspath(file_path))
    worksheet = workbook.Sheets(btn_id)
    sheet_name = worksheet.Name
    for button in buttons:
        button.macro.sheet_name = sheet_name
        button.create(worksheet)
        if workbook:
            module = workbook.VBProject.VBComponents.Add(1)
            module.CodeModule.AddFromString(button.macro.code.strip())
    if workbook:
        workbook.SaveAs(os.path.abspath(new_file_path), FileFormat=52)
        workbook.Close()
