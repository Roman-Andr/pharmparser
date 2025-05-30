import string
from itertools import chain
from typing import List

from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from utils import DataType, Settings
from .base_formatter import BaseFormatter


class DataFormatter(BaseFormatter):
    def __init__(self, settings: Settings, data: DataType, titles: List[str], formatting):
        super().__init__(settings, data, titles)
        self.formatting = formatting

    def format(self, ws: Worksheet):
        offset = 2
        codes = list(self.data.keys())
        grid = [*([] for _ in range(offset)),
                ["Название", self.titles[0]] +
                list(chain(
                    *[[self.titles[i + 1], "Разница"] for i, x in enumerate(self.titles) if x != self.titles[-1]]))]
        names = sorted(list(set(chain(*[list(self.data[x].keys()) for x in self.titles]))), key=lambda k: k.lower())
        for x in names:
            prices = []
            for y in self.titles:
                price1, price2 = self.data[self.titles[0]].get(x, "Нет"), self.data[y].get(x, "Нет")
                prices.append(price2)
                if (price2 == "Нет" or price1 == "Нет") and y != codes[0]:
                    prices.append(0)
                elif y != self.titles[0]:
                    prices.append(float(f"{float(f'{self.formatting(float(price1), float(price2)):.2f}'):+}"))
            row = [x] + prices
            grid.append(row)
        for x in string.ascii_uppercase:
            ws.column_dimensions[x].width = self.settings.cellWidth
        for x in string.ascii_uppercase[3::2]:
            ws.column_dimensions[x].width = self.settings.diffWidth
        ws.column_dimensions["A"].width = self.settings.colWidth

        for x in string.ascii_uppercase[3::2]:
            red_cell, green_cell = PatternFill(bgColor=self.settings.red), PatternFill(bgColor=self.settings.green)
            dxf_red, dxf_green = DifferentialStyle(fill=red_cell), DifferentialStyle(fill=green_cell)
            [ws.conditional_formatting.add(f"{x}{2 + offset}:{x}{len(grid)}", rule) for rule in
             (Rule("cellIs", operator="lessThan", formula=["0"], dxf=dxf_red),
              Rule("cellIs", operator="greaterThan", formula=["0"], dxf=dxf_green))]

        ws.auto_filter.ref = f"A{1 + offset}:{get_column_letter(len(grid[offset]))}{len(grid)}"

        [ws.append(x) for x in grid]
