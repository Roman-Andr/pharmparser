"""End-to-end workbook check that runs without Excel.

The macro buttons need Excel over COM, so they are Windows-only. The openpyxl
half — which produces all of the actual content — is not, and this exercises it:
build every sheet, save a real .xlsx, read it back.
"""

from pathlib import Path

from openpyxl import load_workbook

from pharmparser.config import ExportSettings
from pharmparser.domain import PriceTable
from pharmparser.export import XlsxExporter, write_workbook


def test_workbook_saves_and_reloads(tmp_path: Path, settings: ExportSettings, table: PriceTable) -> None:
    target = write_workbook(settings, table, tmp_path / "data.xlsx")

    reloaded = load_workbook(target)
    assert reloaded.sheetnames == ["Данные", "Проценты", settings.title]

    data = reloaded["Данные"]
    assert [cell.value for cell in data[3]] == [
        "Название",
        "Аптека 1",
        "Аптека 2",
        "Разница",
        "Аптека 3",
        "Разница",
    ]
    assert data["A4"].value == "Аспирин, 100мг"
    assert data["D4"].value == 1.5

    # Percentages differ from absolute differences on the same layout.
    assert reloaded["Проценты"]["D4"].value == 30.0

    analysis = reloaded[settings.title]
    assert analysis["A1"].value == "АНАЛИЗ ЦЕН И АССОРТИМЕНТА"
    assert analysis["F5"].value == 1  # unique items no longer also count as cheapest
    assert analysis["H5"].value == 1


def test_analysis_sheet_is_a_styled_dashboard(tmp_path: Path, settings: ExportSettings, table: PriceTable) -> None:
    sheet = load_workbook(write_workbook(settings, table, tmp_path / "dashboard.xlsx"))[settings.title]

    assert sheet.sheet_view.showGridLines is False
    assert sheet.freeze_panes == "A15"
    assert sheet.auto_filter.ref == "A14:I16"
    assert "A1:I1" in {str(cell_range) for cell_range in sheet.merged_cells.ranges}
    assert sheet["A1"].fill.fgColor.rgb == "0017324D"
    assert sheet["A14"].font.color is not None
    assert sheet["A14"].font.color.rgb == "00FFFFFF"
    assert sheet["H15"].number_format == '0.00 "BYN"'


def test_settings_title_names_the_sheet_and_the_document(tmp_path: Path, table: PriceTable) -> None:
    """B15: the setting used to be validated and then ignored by everything."""
    settings = ExportSettings(title="Сводка")
    reloaded = load_workbook(write_workbook(settings, table, tmp_path / "data.xlsx"))

    assert reloaded.sheetnames == ["Данные", "Проценты", "Сводка"]
    assert reloaded.properties.title == "Сводка"


def test_conditional_formatting_survives_a_save(tmp_path: Path, settings: ExportSettings, table: PriceTable) -> None:
    target = write_workbook(settings, table, tmp_path / "data.xlsx")

    ranges = {str(rng.sqref) for rng in load_workbook(target)["Данные"].conditional_formatting}
    assert ranges == {"D4:D7", "F4:F7"}


def test_autofilter_spans_the_data_block(tmp_path: Path, settings: ExportSettings, table: PriceTable) -> None:
    target = write_workbook(settings, table, tmp_path / "data.xlsx")
    assert load_workbook(target)["Данные"].auto_filter.ref == "A3:F7"


def test_wide_workbook_is_styled_past_column_z(tmp_path: Path, settings: ExportSettings) -> None:
    """Regression for B10, through the writer: 20 pharmacies reach column AN."""
    from pharmparser.domain import Pharmacy

    table = PriceTable.build((Pharmacy(id=str(i), name=f"Аптека {i}"), {"Аспирин": 5.0 + i}) for i in range(20))
    sheet = load_workbook(write_workbook(settings, table, tmp_path / "wide.xlsx"))["Данные"]

    assert sheet.auto_filter.ref == "A3:AN4"
    assert sheet.column_dimensions["AN"].width == settings.diff_width
    assert "AN4" in {str(rng.sqref) for rng in sheet.conditional_formatting}


def test_single_pharmacy_workbook_is_still_valid(tmp_path: Path, settings: ExportSettings) -> None:
    """A profile with one pharmacy has no difference columns at all."""
    from pharmparser.domain import Pharmacy

    table = PriceTable.build([(Pharmacy("1", "Аптека 1"), {"Аспирин": 5.0})])
    reloaded = load_workbook(write_workbook(settings, table, tmp_path / "single.xlsx"))

    assert [cell.value for cell in reloaded["Данные"][3]] == ["Название", "Аптека 1"]
    analysis = reloaded[settings.title]
    assert analysis["D6"].value == 0  # mean competitor assortment
    assert analysis["F5"].value == 0  # no market comparison is possible
    assert analysis["H5"].value == 0


def test_xlsx_exporter_writes_to_the_configured_file_name(
    tmp_path: Path, settings: ExportSettings, table: PriceTable, monkeypatch
) -> None:
    monkeypatch.chdir(tmp_path)
    assert XlsxExporter().export(settings, table) == Path("data.xlsx")
    assert (tmp_path / "data.xlsx").exists()
