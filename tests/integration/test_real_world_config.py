"""Compatibility tests against a real user's config.json.

``tests/fixtures/real_world_config.json`` is a real configuration with the
credentials redacted — six profiles, Cyrillic and Latin pharmacy names, names
carrying trailing and doubled spaces, the same pharmacy id reused across
profiles, a full browser header set (including a lowercase ``host``, an explicit
``Content-Type`` alongside form-encoded data and the browser's whole sixteen-key
cookie) and non-default column widths.

These guard the promise that existing config files keep working unchanged.
"""

import asyncio
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pharmparser.cli import main
from pharmparser.config import load_config, save_config
from pharmparser.export import export_with_macros
from pharmparser.scraping import scrape_profile

from ..endpoint import FakeEndpoint
from ..pages import page as build_page
from ..pages import row

FIXTURE = Path(__file__).parent.parent / "fixtures" / "real_world_config.json"


def page(price: str) -> str:
    return build_page(
        row("Аспирин", "100мг", f"от {price} р.", ls=1),
        row("Цитрамон", "10шт", f"от {price} р.", ls=2),
    )


@pytest.fixture
def config_file(tmp_path: Path, endpoint: FakeEndpoint) -> Path:
    """The real configuration, pointed at the local endpoint instead of tabletka.by."""
    path = tmp_path / "config.json"
    raw = json.loads(FIXTURE.read_text(encoding="utf-8"))
    raw["request"]["url"] = endpoint.url
    path.write_text(json.dumps(raw, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


@pytest.fixture
def served(endpoint: FakeEndpoint, config_file: Path) -> FakeEndpoint:
    """A distinct price per pharmacy, so misalignment would show up in the report."""
    ids = sorted({entry.pharmacy_id for profile in load_config(config_file).profiles for entry in profile.pharmacies})
    for i, pharmacy_id in enumerate(ids):
        endpoint.serve(pharmacy_id, page(f"{5 + i},00"), price_count=2)
    return endpoint


# -- loading -------------------------------------------------------------------


def test_loads_without_error(config_file: Path) -> None:
    config = load_config(config_file)
    assert [p.name for p in config.profiles] == [f"Profile {i}" for i in range(1, 7)]
    assert [len(p.pharmacies) for p in config.profiles] == [9, 5, 3, 3, 3, 8]


def test_custom_widths_are_read(config_file: Path) -> None:
    settings = load_config(config_file).settings
    assert (settings.col_width, settings.cell_width, settings.diff_width) == (45, 13, 9)


def test_awkward_pharmacy_names_are_preserved(config_file: Path) -> None:
    """Trailing spaces, doubled spaces, "№1" and "2-2" are all legal names."""
    profiles = {p.name: p for p in load_config(config_file).profiles}
    assert "Искамед " in [e.name for e in profiles["Profile 2"].pharmacies]
    assert "ADEL  Сов.40А" in [e.name for e in profiles["Profile 2"].pharmacies]
    assert "№1" in [e.name for e in profiles["Profile 6"].pharmacies]
    assert "2-2" in [e.name for e in profiles["Profile 5"].pharmacies]


def test_a_pharmacy_id_may_repeat_across_profiles(config_file: Path) -> None:
    """Pharmacy 381 appears in Profile 2 and Profile 6 under different names."""
    profiles = {p.name: p for p in load_config(config_file).profiles}
    assert "381" in [e.pharmacy_id for e in profiles["Profile 2"].pharmacies]
    assert "381" in [e.pharmacy_id for e in profiles["Profile 6"].pharmacies]


def test_round_trip_is_byte_for_byte_identical(tmp_path: Path) -> None:
    """Saving must not reorder, rename or drop anything (A6).

    Read straight from the fixture rather than the rewritten copy, so the URL the
    real file carries — no trailing slash — is part of what must survive (B17).
    """
    original = tmp_path / "original.json"
    original.write_text(FIXTURE.read_text(encoding="utf-8"), encoding="utf-8")
    out = tmp_path / "saved.json"
    save_config(load_config(original), out)
    assert json.loads(out.read_text(encoding="utf-8")) == json.loads(
        FIXTURE.read_text(encoding="utf-8")
    )


# -- scraping ------------------------------------------------------------------


def test_browser_headers_survive_to_the_request(config_file: Path, tmp_path: Path, served: FakeEndpoint) -> None:
    """An explicit Content-Type alongside form-encoded data must not be rejected."""
    assert main(["--config", str(config_file), "--output", str(tmp_path / "r.xlsx")]) == 0

    sent = served.requests[0]
    assert sent.headers["Content-Type"] == "application/x-www-form-urlencoded; charset=UTF-8"
    assert sent.headers["host"] == "tabletka.by"
    assert "lim-result=5000" in sent.cookie
    assert {request.pharmacy_id for request in served.requests} >= {"3563"}


# -- exporting -----------------------------------------------------------------


def test_nine_pharmacy_profile_exports(config_file: Path, tmp_path: Path, served: FakeEndpoint) -> None:
    output = tmp_path / "report.xlsx"
    assert main(["--config", str(config_file), "--output", str(output)]) == 0

    sheet = load_workbook(output)["Данные"]
    assert sheet.max_column == 2 + 2 * 8
    assert [cell.value for cell in sheet[3]][:4] == [
        "Название",
        "Аптека 195/3",
        "Добрые леки",
        "Разница",
    ]
    assert sheet.column_dimensions["A"].width == 45
    assert sheet.column_dimensions["C"].width == 13
    assert sheet.column_dimensions["D"].width == 9


def test_every_profile_can_be_selected(config_file: Path, tmp_path: Path, served: FakeEndpoint) -> None:
    counts = {"Profile 1": 9, "Profile 2": 5, "Profile 3": 3, "Profile 4": 3, "Profile 5": 3, "Profile 6": 8}
    for name, pharmacies in counts.items():
        output = tmp_path / f"{name}.xlsx"
        assert main(["--config", str(config_file), "--profile", name, "--output", str(output)]) == 0
        assert load_workbook(output)["Данные"].max_column == 2 + 2 * (pharmacies - 1)


def test_the_whole_browser_cookie_survives_to_the_request(
    config_file: Path, tmp_path: Path, served: FakeEndpoint
) -> None:
    """Only ``lim-result`` is rewritten; the other fifteen cookies travel untouched."""
    assert main(["--config", str(config_file), "--output", str(tmp_path / "r.xlsx")]) == 0

    configured = load_config(config_file).request.cookie
    sent = served.requests[0].cookie
    keys = [cookie.split("=", 1)[0].strip() for cookie in sent.split(";")]
    assert keys == [cookie.split("=", 1)[0].strip() for cookie in configured.split(";")]
    assert "region=%D0%93%D0%BE%D0%BC%D0%B5%D0%BB%D1%8C" in sent


def test_the_configured_title_names_the_analysis_sheet(
    config_file: Path, tmp_path: Path, served: FakeEndpoint
) -> None:
    """B15, against the real value: this config sets ``title`` to "Анализ"."""
    output = tmp_path / "report.xlsx"
    assert main(["--config", str(config_file), "--profile", "Profile 1", "--output", str(output)]) == 0

    workbook = load_workbook(output)
    assert load_config(config_file).settings.title == "Анализ"
    assert workbook.sheetnames == ["Данные", "Проценты", "Анализ"]
    assert workbook.properties.title == "Анализ"


def test_the_macro_export_runs_over_the_real_profile(
    config_file: Path, tmp_path: Path, served: FakeEndpoint, excel_sessions: list
) -> None:
    """The nine-pharmacy profile through the .xlsm path: one Excel, one output file."""
    config = load_config(config_file)
    table = asyncio.run(scrape_profile(config.request, config.profiles[0].pharmacies))

    target = export_with_macros(config.settings, table, tmp_path / "data.xlsm", use_excel=True)

    assert len(excel_sessions) == 1
    assert sorted(path.name for path in tmp_path.iterdir()) == ["config.json", "data.xlsm"]
    workbook = excel_sessions[0].opened[0]
    # 8 competitors -> 8 difference columns -> 16 sort buttons plus the 2 filter ones.
    assert len(workbook.Sheets("Данные").Shapes.shapes) == 18
    assert target.exists()
