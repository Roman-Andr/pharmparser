from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from pharmparser.application import HistoryRepository, ReportService
from pharmparser.application.models import ProfileRecord
from pharmparser.application.settings import DesktopSettings
from pharmparser.domain import Product, RunStatus


def test_five_sheet_xlsm_report_uses_tables_and_partial_warning(tmp_path: Path) -> None:
    repository = HistoryRepository(tmp_path / "history.sqlite3")
    profile = ProfileRecord.model_validate(
        {
            "name": "Тест",
            "reference_pharmacy_id": "1",
            "pharmacies": [
                {"id": "1", "name": "Наша", "url": "https://tabletka.by/pharmacies/1"},
                {"id": "2", "name": "Конкурент", "url": "https://tabletka.by/pharmacies/2"},
                {"id": "3", "name": "Ошибка", "url": "https://tabletka.by/pharmacies/3"},
            ],
        }
    )
    run_id = repository.create_run(profile)
    product = Product("Аспирин", "100 мг", "Bayer")
    for pharmacy_id, name, amount in (("1", "Наша", "5.00"), ("2", "Конкурент", "4.50")):
        repository.start_attempt(run_id, pharmacy_id, name)
        repository.store_prices(run_id, pharmacy_id, [(product, Decimal(amount))])
        repository.finish_attempt(run_id, pharmacy_id, status="completed", items=1)
    repository.start_attempt(run_id, "3", "Ошибка")
    repository.finish_attempt(run_id, "3", status="failed", error_code="timeout", error_message="timeout")
    repository.add_warning(run_id, "pharmacy_failed", "Ошибка: timeout", "3")
    repository.set_status(run_id, RunStatus.PARTIAL)

    target = ReportService(repository).export(
        run_id,
        DesktopSettings(output_directory=str(tmp_path), report_format="xlsm"),
    )
    workbook = load_workbook(target, keep_vba=True)
    assert workbook.sheetnames == ["Обзор", "Сравнение", "Изменения", "История", "Проблемы"]
    assert "неполный" in workbook["Обзор"]["A2"].value.lower()
    assert "ComparisonTable" in workbook["Сравнение"].tables
    assert "ChangesTable" in workbook["Изменения"].tables
    assert workbook.vba_archive is not None
    assert workbook["Проблемы"]["C2"].value == "pharmacy_failed"
