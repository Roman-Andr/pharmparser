"""End-to-end run of the CLI against a real price endpoint on localhost."""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pharmparser.cli import main
from pharmparser.platform_ import supports_excel_macros

from ..endpoint import FakeEndpoint
from ..pages import page, row


def PAGE(price: str) -> str:
    return page(row("Аспирин", "100мг", f"от {price} р."))


@pytest.fixture
def config_file(tmp_path: Path, endpoint: FakeEndpoint) -> Path:
    path = tmp_path / "config.json"
    path.write_text(
        json.dumps(
            {
                "profiles": {
                    "Основной": {
                        "Аптека 1": "https://tabletka.by/pharmacies/111",
                        "Аптека 2": "https://tabletka.by/pharmacies/222",
                    },
                    "Пустой": {},
                },
                "settings": {"fileName": "out.xlsx", "title": "Тест"},
                "request": {
                    "url": endpoint.url,
                    "headers": {"Cookie": "PHPSESSID=abc; lim-result=5000"},
                    "data": {"_csrf": "token"},
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return path


@pytest.fixture
def two_pharmacies(endpoint: FakeEndpoint) -> FakeEndpoint:
    """Аптека 1 at 5.00, Аптека 2 at 6.50 — a 1.50 difference."""
    endpoint.serve("111", PAGE("5,00"))
    endpoint.serve("222", PAGE("6,50"))
    return endpoint


def test_cli_writes_a_workbook(config_file: Path, tmp_path: Path, two_pharmacies: FakeEndpoint) -> None:
    output = tmp_path / "report.xlsx"
    assert main(["--config", str(config_file), "--output", str(output)]) == 0

    sheet = load_workbook(output)["Данные"]
    assert [cell.value for cell in sheet[3]] == ["Название", "Аптека 1", "Аптека 2", "Разница"]
    assert [cell.value for cell in sheet[4]] == ["Аспирин, 100мг", 5.0, 6.5, 1.5]


def test_cli_defaults_to_the_first_profile(config_file: Path, tmp_path: Path, two_pharmacies: FakeEndpoint) -> None:
    output = tmp_path / "report.xlsx"
    assert main(["--config", str(config_file), "--output", str(output)]) == 0
    assert output.exists()


def test_cli_reports_an_unknown_profile(config_file: Path, capsys: pytest.CaptureFixture) -> None:
    assert main(["--config", str(config_file), "--profile", "Нет такого"]) == 1
    assert "available profiles: Основной, Пустой" in capsys.readouterr().err


def test_cli_reports_an_empty_profile(config_file: Path, capsys: pytest.CaptureFixture) -> None:
    """B4: an empty profile used to crash inside a worker thread with Pool(0)."""
    assert main(["--config", str(config_file), "--profile", "Пустой"]) == 1
    assert "no pharmacies" in capsys.readouterr().err


def test_cli_reports_a_missing_config(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    assert main(["--config", str(tmp_path / "nope.json")]) == 1
    assert "not found" in capsys.readouterr().err


def test_cli_reports_a_scrape_failure(
    config_file: Path, tmp_path: Path, endpoint: FakeEndpoint, capsys: pytest.CaptureFixture
) -> None:
    """B6: scrape failures reach the user instead of vanishing."""
    endpoint.fail(500)
    assert main(["--config", str(config_file), "--output", str(tmp_path / "x.xlsx")]) == 1
    assert "could not be parsed" in capsys.readouterr().err


def test_env_cookie_reaches_the_request(
    config_file: Path, tmp_path: Path, two_pharmacies: FakeEndpoint, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("PHARMPARSER_COOKIE", "PHPSESSID=from-env; lim-result=5000")
    assert main(["--config", str(config_file), "--output", str(tmp_path / "r.xlsx")]) == 0
    assert all("from-env" in request.cookie for request in two_pharmacies.requests)


@pytest.mark.skipif(supports_excel_macros(), reason="Excel is available, so macros are not skipped")
def test_macros_flag_degrades_gracefully_off_windows(
    config_file: Path, tmp_path: Path, two_pharmacies: FakeEndpoint, caplog: pytest.LogCaptureFixture
) -> None:
    output = tmp_path / "report.xlsx"
    with caplog.at_level("WARNING"):
        assert main(["--config", str(config_file), "--output", str(output), "--macros"]) == 0
    assert output.exists()
    assert "Windows only" in caplog.text


def test_cache_flag_reuses_a_cached_run(
    config_file: Path, tmp_path: Path, two_pharmacies: FakeEndpoint, monkeypatch: pytest.MonkeyPatch
) -> None:
    """--cache makes the second run answer from disk without touching the network."""
    monkeypatch.chdir(tmp_path)  # caches land beside the config, not in the repo
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"

    assert main(["--config", str(config_file), "--output", str(first), "--cache"]) == 0
    requests_so_far = len(two_pharmacies.requests)

    assert main(["--config", str(config_file), "--output", str(second), "--cache"]) == 0
    assert len(two_pharmacies.requests) == requests_so_far, "the second run went to the network"
    assert load_workbook(second).sheetnames == load_workbook(first).sheetnames


def test_without_the_cache_flag_nothing_is_written_to_the_cache(
    config_file: Path, tmp_path: Path, two_pharmacies: FakeEndpoint, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.chdir(tmp_path)
    assert main(["--config", str(config_file), "--output", str(tmp_path / "r.xlsx")]) == 0
    assert not list(tmp_path.glob(".cache-*.json"))
