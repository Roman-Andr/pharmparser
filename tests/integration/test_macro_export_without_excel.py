"""The default .xlsm path: no Excel, no COM, no Windows.

This is the test that would have been impossible before — it builds the real
deliverable on the CI Linux box and then takes it apart again.
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path

import pytest
from oletools.olevba import VBA_Parser

from pharmparser.config import ExportSettings
from pharmparser.domain import PriceTable
from pharmparser.export import MacroExporter, export_with_macros

VBA_PART = "xl/vbaProject.bin"
FMLA_MACRO = re.compile(r"<x:FmlaMacro>\[0\]!(\w+)</x:FmlaMacro>")


@pytest.fixture
def report(tmp_path: Path, settings: ExportSettings, table: PriceTable) -> Path:
    return export_with_macros(settings, table, tmp_path / "data.xlsm")


def parts(path: Path) -> list[str]:
    return sorted(zipfile.ZipFile(path).namelist())


def vml(path: Path) -> str:
    archive = zipfile.ZipFile(path)
    return "".join(archive.read(name).decode() for name in archive.namelist() if "vmlDrawing" in name)


def test_the_workbook_is_a_macro_enabled_package(report: Path) -> None:
    assert VBA_PART in parts(report)
    content_types = zipfile.ZipFile(report).read("[Content_Types].xml").decode()
    assert "application/vnd.ms-excel.sheet.macroEnabled.main+xml" in content_types
    assert 'Extension="bin"' in content_types
    assert 'Extension="vml"' in content_types


def test_the_vba_project_relationship_is_declared(report: Path) -> None:
    rels = zipfile.ZipFile(report).read("xl/_rels/workbook.xml.rels").decode()
    assert "vbaProject" in rels and 'Target="vbaProject.bin"' in rels


def test_buttons_are_drawn_only_on_the_data_sheets(report: Path) -> None:
    drawings = [name for name in parts(report) if "vmlDrawing" in name]
    assert len(drawings) == 2, "Данные and Проценты, not Анализ"
    # Apply + Remove, then an up/down pair per difference column (D and F here).
    assert vml(report).count("<x:FmlaMacro>") == 2 * (2 + 4)


def test_each_sheet_links_its_legacy_drawing(report: Path) -> None:
    archive = zipfile.ZipFile(report)
    for index in (1, 2):
        rels = archive.read(f"xl/worksheets/_rels/sheet{index}.xml.rels").decode()
        assert "vmlDrawing" in rels
        assert "<legacyDrawing" in archive.read(f"xl/worksheets/sheet{index}.xml").decode()


def test_every_button_points_at_a_macro_that_exists(report: Path, tmp_path: Path) -> None:
    """The failure this rules out is a button wired to a Sub that was never compiled."""
    referenced = set(FMLA_MACRO.findall(vml(report)))
    assert referenced, "no buttons were bound"

    blob = tmp_path / "vbaProject.bin"
    blob.write_bytes(zipfile.ZipFile(report).read(VBA_PART))
    compiled = {
        name
        for _, _, module, code in VBA_Parser(str(blob)).extract_macros()
        if "PharmParser" in module
        for name in re.findall(r"Sub (\w+)\(", code)
    }
    assert referenced <= compiled, f"buttons reference missing macros: {referenced - compiled}"


def test_the_compiled_module_does_not_change_legacy_button_placement(
    report: Path, tmp_path: Path
) -> None:
    """VML form buttons can reject ``Shape.Placement`` in Excel with error 80020009.

    Their VML anchors already live above the filtered range and do not opt into
    moving or sizing with cells, so the generated VBA must not mutate Placement.
    """
    blob = tmp_path / "vbaProject.bin"
    blob.write_bytes(zipfile.ZipFile(report).read(VBA_PART))
    for _, _, module, code in VBA_Parser(str(blob)).extract_macros():
        if "PharmParser" in module:
            assert code.isascii(), "module streams are stored in the project code page"
            assert ".Placement" not in code
            assert "PharmParserPinShapes" not in code


def test_the_plain_sheet_data_survives_packaging(report: Path, settings: ExportSettings) -> None:
    """Also proves the patched worksheet XML is still well-formed."""
    from openpyxl import load_workbook

    workbook = load_workbook(report, keep_vba=True)
    assert workbook.sheetnames == ["Данные", "Проценты", settings.title]
    assert next(cell.value for cell in workbook["Данные"][3]) == "Название"


def test_exporter_leaves_no_temp_files(tmp_path: Path, settings: ExportSettings, table: PriceTable) -> None:
    target = tmp_path / "data.xlsm"
    assert MacroExporter().export(settings, table, target) == target
    assert [path.name for path in tmp_path.iterdir()] == ["data.xlsm"]


def test_it_replaces_a_stale_report_in_place(tmp_path: Path, settings: ExportSettings, table: PriceTable) -> None:
    target = tmp_path / "data.xlsm"
    target.write_bytes(b"stale")
    export_with_macros(settings, table, target)
    assert zipfile.is_zipfile(target)
