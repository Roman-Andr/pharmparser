"""The Controller, exercised without a display.

That it can be exercised at all is the point of phase 5: this logic used to live
inside a CTk subclass, so none of it could run without a display server (A1).
"""

from __future__ import annotations

import json
from collections.abc import Sequence
from pathlib import Path

import pytest

from pharmparser.cache import write_table
from pharmparser.config import AppConfig, ConfigError, PharmacyEntry, Profile, RequestConfig
from pharmparser.controller import Controller
from pharmparser.domain import Pharmacy, PriceTable
from pharmparser.export import write_workbook
from pharmparser.scraping import ScrapeError

RAW = {
    "profiles": {
        "Основной": {
            "Аптека 1": "https://tabletka.by/pharmacies/111",
            "Аптека 2": "https://tabletka.by/pharmacies/222",
        }
    },
    "settings": {"fileName": "out.xlsx", "title": "Сводка"},
    "request": {
        "url": "https://tabletka.by/ajax-request/reload-pharmacy-price",
        "headers": {"Cookie": "PHPSESSID=abc; lim-result=5000"},
        "data": {"_csrf": "token"},
    },
}


class RecordingExporter:
    """An Exporter that writes a real workbook and remembers what it was given."""

    def __init__(self, directory: Path) -> None:
        self.directory = directory
        self.tables: list[PriceTable] = []

    def default_path(self, settings) -> Path:
        return self.directory / settings.file_name

    def export(self, settings, table: PriceTable, path: Path | None = None) -> Path:
        self.tables.append(table)
        return write_workbook(settings, table, path or self.default_path(settings))


def scraped(names: Sequence[str]) -> PriceTable:
    return PriceTable.build(
        (Pharmacy(id=str(i), name=name), {"Аспирин": 5.0 + i}) for i, name in enumerate(names, start=1)
    )


@pytest.fixture
def config_file(tmp_path: Path) -> Path:
    path = tmp_path / "config.json"
    path.write_text(json.dumps(RAW, ensure_ascii=False), encoding="utf-8")
    return path


@pytest.fixture
def controller(config_file: Path, tmp_path: Path) -> Controller:
    calls: list[tuple] = []

    def fake_scrape(request: RequestConfig, pharmacies) -> PriceTable:
        calls.append((request, tuple(pharmacies)))
        return scraped([entry.name for entry in pharmacies if entry.is_complete])

    made = Controller.load(
        config_file,
        exporter=RecordingExporter(tmp_path),
        scrape=fake_scrape,
        cache_file=lambda name: tmp_path / f"cache-{name}.json",
    )
    made.scrape_calls = calls  # type: ignore[attr-defined]
    return made


# -- configuration -------------------------------------------------------------


def test_load_reads_the_configuration(controller: Controller) -> None:
    assert [profile.name for profile in controller.profiles] == ["Основной"]
    assert controller.settings.title == "Сводка"
    assert controller.request.cookie.startswith("PHPSESSID=")


def test_load_reports_a_missing_file_as_a_config_error(tmp_path: Path) -> None:
    with pytest.raises(ConfigError):
        Controller.load(tmp_path / "nope.json")


def test_an_empty_configuration_still_offers_one_profile(tmp_path: Path) -> None:
    config = AppConfig(request=RequestConfig(headers={"Cookie": "x=1"}))
    assert [p.name for p in Controller(config, tmp_path / "c.json").profiles] == ["Profile 1"]


def test_save_keeps_the_names_the_user_chose(controller: Controller, config_file: Path) -> None:
    """A6: profile names used to be regenerated as "Profile 1..N" on every save."""
    controller.save([Profile(name="Мой профиль", pharmacies=[])])
    assert list(json.loads(config_file.read_text(encoding="utf-8"))["profiles"]) == ["Мой профиль"]


def test_save_leaves_the_rest_of_the_configuration_alone(controller: Controller, config_file: Path) -> None:
    controller.save(controller.profiles)
    saved = json.loads(config_file.read_text(encoding="utf-8"))
    assert saved["settings"]["title"] == "Сводка"
    assert saved["request"]["headers"]["Cookie"].startswith("PHPSESSID=")


@pytest.mark.parametrize(
    ("existing", "expected"),
    [
        ([], "Profile 1"),
        (["Profile 1"], "Profile 2"),
        (["Profile 1", "Profile 2"], "Profile 3"),
        (["Мой", "Profile 2"], "Profile 3"),
        (["Profile 1", "Profile 3"], "Profile 4"),
    ],
)
def test_next_profile_name_avoids_collisions(existing: list[str], expected: str) -> None:
    assert Controller.next_profile_name(existing) == expected


# -- collecting ----------------------------------------------------------------


def test_collect_scrapes(controller: Controller) -> None:
    table = controller.collect(controller.profiles[0])
    assert [p.name for p in table.pharmacies] == ["Аптека 1", "Аптека 2"]


def test_the_cache_is_entirely_opt_in(controller: Controller, tmp_path: Path) -> None:
    """A7: the old code wrote a shared data.json on every run whatever was asked."""
    controller.collect(controller.profiles[0])
    assert not (tmp_path / "cache-Основной.json").exists()


def test_collect_writes_the_cache_when_asked(controller: Controller, tmp_path: Path) -> None:
    controller.collect(controller.profiles[0], use_cache=True)
    assert (tmp_path / "cache-Основной.json").exists()


def test_collect_ignores_an_existing_cache_unless_asked(controller: Controller, tmp_path: Path) -> None:
    write_table(scraped(["Из кэша"]), tmp_path / "cache-Основной.json")
    table = controller.collect(controller.profiles[0], use_cache=False)
    assert [p.name for p in table.pharmacies] == ["Аптека 1", "Аптека 2"]


def test_collect_uses_the_cache_when_asked(controller: Controller, tmp_path: Path) -> None:
    write_table(scraped(["Из кэша"]), tmp_path / "cache-Основной.json")
    table = controller.collect(controller.profiles[0], use_cache=True)
    assert [p.name for p in table.pharmacies] == ["Из кэша"]
    assert controller.scrape_calls == []  # type: ignore[attr-defined]


def test_caches_are_per_profile(controller: Controller, tmp_path: Path) -> None:
    """A7: one shared data.json meant switching profile silently reused its prices."""
    write_table(scraped(["Чужой кэш"]), tmp_path / "cache-Другой.json")
    table = controller.collect(controller.profiles[0], use_cache=True)
    assert [p.name for p in table.pharmacies] == ["Аптека 1", "Аптека 2"]


def test_an_unreadable_cache_falls_back_to_scraping(
    controller: Controller, tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    (tmp_path / "cache-Основной.json").write_text("{ not json", encoding="utf-8")
    with caplog.at_level("WARNING"):
        table = controller.collect(controller.profiles[0], use_cache=True)
    assert [p.name for p in table.pharmacies] == ["Аптека 1", "Аптека 2"]
    assert "unreadable cache" in caplog.text


def test_a_scrape_failure_reaches_the_caller(config_file: Path, tmp_path: Path) -> None:
    """B6: failures used to be collected into a list nothing ever appended to."""

    def failing(request, pharmacies):
        raise ScrapeError("tabletka.by said no")

    controller = Controller.load(
        config_file, scrape=failing, cache_file=lambda name: tmp_path / f"{name}.json"
    )
    with pytest.raises(ScrapeError, match=r"tabletka\.by said no"):
        controller.collect(controller.profiles[0])


# -- exporting and the whole run ------------------------------------------------


def test_export_writes_the_configured_file(controller: Controller, tmp_path: Path) -> None:
    path = controller.export(scraped(["Аптека 1", "Аптека 2"]))
    assert path == (tmp_path / "out.xlsx").absolute()
    assert path.exists()


def test_run_collects_then_exports(controller: Controller, tmp_path: Path) -> None:
    from openpyxl import load_workbook

    path = controller.run(controller.profiles[0])
    assert load_workbook(path).sheetnames == ["Данные", "Проценты", "Сводка"]
    assert len(controller.scrape_calls) == 1  # type: ignore[attr-defined]


def test_run_passes_the_profiles_pharmacies_to_the_scraper(controller: Controller) -> None:
    controller.run(controller.profiles[0])
    _, pharmacies = controller.scrape_calls[0]  # type: ignore[attr-defined]
    assert [entry.name for entry in pharmacies] == ["Аптека 1", "Аптека 2"]
    assert all(isinstance(entry, PharmacyEntry) for entry in pharmacies)


# -- profile selection ---------------------------------------------------------


def test_select_profile_defaults_to_the_first(controller: Controller) -> None:
    assert controller.select_profile(None).name == "Основной"


def test_select_profile_finds_by_name(controller: Controller) -> None:
    assert controller.select_profile("Основной").name == "Основной"


def test_select_profile_lists_what_is_available(controller: Controller) -> None:
    with pytest.raises(ConfigError, match="available profiles: Основной"):
        controller.select_profile("Нет такого")


def test_select_profile_reports_an_empty_configuration(tmp_path: Path) -> None:
    config = AppConfig(request=RequestConfig(headers={"Cookie": "x=1"}))
    with pytest.raises(ConfigError, match="no profiles are configured"):
        Controller(config, tmp_path / "c.json").select_profile(None)
