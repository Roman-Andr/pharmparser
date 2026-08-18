"""Sheet-level tests for DataFormatter."""

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from pharmparser.domain import Pharmacy, PriceTable, absolute_difference, percentage_difference
from pharmparser.excel.formatters import DataFormatter
from pharmparser.excel.formatters.data_formatter import HEADER_OFFSET
from pharmparser.utils import Settings


def build_rows(settings: Settings, table: PriceTable, difference) -> list[list]:
    ws = Workbook().active
    DataFormatter(settings, table, difference).format(ws)
    return [list(row) for row in ws.iter_rows(values_only=True)]


def test_header_row_alternates_pharmacy_and_difference(settings: Settings, table: PriceTable) -> None:
    rows = build_rows(settings, table, absolute_difference)
    assert rows[HEADER_OFFSET] == ["Название", "Аптека 1", "Аптека 2", "Разница", "Аптека 3", "Разница"]


def test_rows_are_sorted_case_insensitively(settings: Settings, table: PriceTable) -> None:
    rows = build_rows(settings, table, absolute_difference)
    assert [row[0] for row in rows[3:]] == [
        "Аспирин, 100мг",
        "Ибупрофен, 200мг",
        "Парацетамол, 500мг",
        "Цитрамон, 10шт",
    ]


def test_differences_are_signed_against_the_reference_pharmacy(
    settings: Settings, table: PriceTable
) -> None:
    rows = build_rows(settings, table, absolute_difference)
    assert rows[3] == ["Аспирин, 100мг", 5.0, 6.5, 1.5, 7.0, 2.0]
    assert rows[5][:4] == ["Парацетамол, 500мг", 3.0, 2.5, -0.5]


def test_percentage_formatting_is_relative(settings: Settings, table: PriceTable) -> None:
    rows = build_rows(settings, table, percentage_difference)
    assert rows[3][3] == 30.0


def test_missing_items_render_as_a_label_with_a_blank_difference(
    settings: Settings, table: PriceTable
) -> None:
    """Regression for B9.

    A pharmacy that does not stock an item shows "Нет", and its difference column is
    left empty rather than 0 — a 0 would be indistinguishable from prices that
    genuinely match.
    """
    rows = build_rows(settings, table, absolute_difference)
    assert rows[4] == ["Ибупрофен, 200мг", "Нет", "Нет", None, 4.0, None]
    assert rows[6] == ["Цитрамон, 10шт", 2.0, "Нет", None, "Нет", None]


def test_autofilter_spans_the_data_block(settings: Settings, table: PriceTable) -> None:
    ws = Workbook().active
    DataFormatter(settings, table, absolute_difference).format(ws)
    assert ws.auto_filter.ref == "A3:F7"


def wide_table(pharmacy_count: int) -> PriceTable:
    return PriceTable.build(
        (Pharmacy(id=str(i), name=f"Аптека {i}"), {"Аспирин": 5.0 + i}) for i in range(pharmacy_count)
    )


@pytest.mark.parametrize("pharmacy_count", [3, 13, 20])
def test_difference_columns_are_styled_beyond_column_z(pharmacy_count: int, settings: Settings) -> None:
    """Regression for B10.

    The formatters used to walk ``string.ascii_uppercase``, so column widths and the
    conditional formatting stopped at Z — i.e. at roughly 13 pharmacies.
    """
    ws = Workbook().active
    DataFormatter(settings, wide_table(pharmacy_count), absolute_difference).format(ws)

    total_columns = 2 + 2 * (pharmacy_count - 1)
    for column in range(4, total_columns + 1, 2):
        letter = get_column_letter(column)
        assert ws.column_dimensions[letter].width == settings.diffWidth, letter

    styled = {str(rng.sqref).split(":")[0].rstrip("0123456789") for rng in ws.conditional_formatting}
    assert get_column_letter(total_columns) in styled


def test_wide_table_autofilter_uses_the_real_last_column(settings: Settings) -> None:
    ws = Workbook().active
    DataFormatter(settings, wide_table(20), absolute_difference).format(ws)
    # 20 pharmacies -> 40 columns -> AN.
    assert ws.auto_filter.ref == "A3:AN4"
