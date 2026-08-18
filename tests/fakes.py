"""A stand-in for the Excel COM objects the VBA injector drives.

Enough of the object model to record what the injector asks Excel to do, so the
Windows-only export path can be exercised on any platform. Attribute names match
Excel's, capitals included, because that is what the code under test calls.
"""

from __future__ import annotations

import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


class FakeCharacters:
    def __init__(self, frame: FakeTextFrame) -> None:
        self._frame = frame

    @property
    def Text(self) -> str:
        return self._frame.text

    @Text.setter
    def Text(self, value: str) -> None:
        self._frame.text = value


class FakeTextFrame:
    def __init__(self) -> None:
        self.text = ""
        self.HorizontalAlignment: int | None = None
        self.VerticalAlignment: int | None = None

    def Characters(self) -> FakeCharacters:
        return FakeCharacters(self)


@dataclass
class FakeColour:
    RGB: int | None = None


@dataclass
class FakeFill:
    BackColor: FakeColour = field(default_factory=FakeColour)
    ForeColor: FakeColour = field(default_factory=FakeColour)


@dataclass
class FakeShape:
    Name: str
    Left: float
    Top: float
    Width: float
    Height: float
    OnAction: str | None = None
    TextFrame: FakeTextFrame = field(default_factory=FakeTextFrame)
    Fill: FakeFill = field(default_factory=FakeFill)


class FakeShapes:
    def __init__(self) -> None:
        self.shapes: list[FakeShape] = []

    def AddShape(self, _kind: int, left: float, top: float, width: float, height: float) -> FakeShape:
        shape = FakeShape(f"Shape {len(self.shapes) + 1}", left, top, width, height)
        self.shapes.append(shape)
        return shape


@dataclass
class FakeCell:
    Left: float
    Top: float
    Width: float
    Height: float


class FakeWorksheet:
    def __init__(self, name: str) -> None:
        self.Name = name
        self.Shapes = FakeShapes()
        self.requested_ranges: list[str] = []

    def Range(self, address: str) -> FakeCell:
        self.requested_ranges.append(address)
        return FakeCell(Left=10.0, Top=20.0, Width=64.0, Height=18.0)


class FakeCodeModule:
    def __init__(self) -> None:
        self.sources: list[str] = []

    def AddFromString(self, code: str) -> None:
        self.sources.append(code)


class FakeVBComponent:
    def __init__(self) -> None:
        self.CodeModule = FakeCodeModule()


class FakeVBComponents:
    def __init__(self) -> None:
        self.components: list[FakeVBComponent] = []

    def Add(self, _kind: int) -> FakeVBComponent:
        component = FakeVBComponent()
        self.components.append(component)
        return component


class FakeVBProject:
    def __init__(self) -> None:
        self.VBComponents = FakeVBComponents()


class FakeWorkbook:
    def __init__(self, path: Path, sheet_names: list[str]) -> None:
        self.FullName = str(path)
        self._path = path
        self._sheets = {name: FakeWorksheet(name) for name in sheet_names}
        self.VBProject = FakeVBProject()
        self.saved_as: Path | None = None
        self.closed = False

    def Sheets(self, name: str) -> FakeWorksheet:
        return self._sheets[name]

    def SaveAs(self, path: str, FileFormat: int) -> None:
        self.file_format = FileFormat
        self.saved_as = Path(path)
        shutil.copyfile(self._path, self.saved_as)

    def Close(self, SaveChanges: bool = True) -> None:
        self.closed = True


class FakeWorkbooks:
    def __init__(self, excel: FakeExcel) -> None:
        self._excel = excel

    def __iter__(self) -> Any:
        return iter(self._excel.open_workbooks)

    def Open(self, path: str) -> FakeWorkbook:
        from openpyxl import load_workbook

        workbook = FakeWorkbook(Path(path), load_workbook(path).sheetnames)
        self._excel.opened.append(workbook)
        self._excel.open_workbooks.append(workbook)
        return workbook


class FakeExcel:
    """One Excel process. ``sessions`` on the factory counts how many were started."""

    def __init__(self) -> None:
        self.Visible: bool | None = None
        self.DisplayAlerts: bool | None = None
        self.opened: list[FakeWorkbook] = []
        self.open_workbooks: list[FakeWorkbook] = []
        self.quit = False
        self.Workbooks = FakeWorkbooks(self)

    def Quit(self) -> None:
        self.quit = True
